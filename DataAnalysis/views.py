import random

from django.shortcuts import get_object_or_404, render

# Create your views here.
# views.py
import os , sys
import sqlite3
from django.shortcuts import render, redirect
from django.conf import settings
import jdatetime
import numpy as np
import xlrd
import xlwt

from otp_manager.models import OTPVar_Enum, SMS_Recievers, SMS_Template, SMSServiceTemplate_Enum
from otp_manager.service import send_sms
from .folder_utils.sepidar_date import format_jalali_date, format_jalali_datetime, havale_format_jalali_datetime
from .forms import DBUploadForm
from .models import InvoiceItem, Sale,SMSLog
from persiantools.jdatetime import JalaliDate
from user_management.utils import check_server
from django.utils import timezone


SERVER = check_server()


def upload_db(request):
    if request.method == "POST":
        form = DBUploadForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded_file = form.cleaned_data["file"]
            temp_path = os.path.join(settings.MEDIA_ROOT, uploaded_file.name)
            with open(temp_path, "wb+") as dest:
                for chunk in uploaded_file.chunks():
                    dest.write(chunk)

            # Connect to uploaded DB
            conn = sqlite3.connect(temp_path)
            cursor = conn.cursor()

            try:
                cursor.execute("SELECT factnum, dat, total, kname, tel, adress FROM facts")  # adjust table name
                rows = cursor.fetchall()

                # Import data
                for row in rows:


                    jalali_str = row[1]  # example: "03/10/27"

                    # Clean and parse Jalali date (assuming format: YY/MM/DD or YYYY/MM/DD)
                    try:
                        parts = jalali_str.replace("“", "").replace("”", "").split("/")
                        if len(parts[0]) == 2:
                            # If year is short like 03 → convert to 1403
                            year = int(parts[0]) + 1400
                        else:
                            year = int(parts[0])
                        month = int(parts[1])
                        day = int(parts[2])

                        gregorian_date = JalaliDate(year, month, day).to_gregorian()
                    except Exception as e:
                        #print(f"⚠️ Error converting date {jalali_str}: {e}")
                        continue  # skip bad records








                    Sale.objects.get_or_create(
                        factnum=row[0],
                        defaults={
                            'dat': gregorian_date,
                            'total': row[2],
                            'kname': row[3],
                            'tel': row[4],
                            'address': row[5],
                       
                        }
                    )
                conn.close()
                os.remove(temp_path)
                return render(request, "UploadDB/upload_success.html", {"count": len(rows)})

            except Exception as e:
                conn.close()
                return render(request, "UploadDB/upload_error.html", {"error": str(e)})

    else:
        form = DBUploadForm()

    return render(request, "UploadDB/upload_db.html", {"form": form})


from django.shortcuts import render
from django.db.models import Sum
from django.db.models.functions import TruncDate

from django.shortcuts import render
from django.db.models import Sum
from django.db.models.functions import TruncDate

from .models import Invoice, InvoiceItem


from django.db.models import Sum
from django.db.models import Sum, F, ExpressionWrapper, DurationField, DateField
from django.db.models.functions import TruncDate
from django.db import models
from datetime import timedelta, time
from django.db.models import Sum, F, DateTimeField, ExpressionWrapper


def dashboard(request):
    top_customers = list(
        Invoice.objects.values('phone')
        .annotate(total_spent=Sum('total_price'))
        .order_by('-total_spent')[:10]
    )

    daily_sales = list(
        Invoice.objects
        .annotate(
            shifted_time=ExpressionWrapper(
                F('created_at') - timedelta(hours=3),
                output_field=DateTimeField()
            )
        )
        .annotate(day=TruncDate('shifted_time'))
        .values('day')
        .annotate(total_day=Sum('total_price'))
        .order_by('day')
    )

    summary = Invoice.objects.aggregate(total_revenue=Sum('total_price'))

    daily_item_volume = list(
        InvoiceItem.objects
        .annotate(
            shifted_time=ExpressionWrapper(
                F('invoice__created_at') - timedelta(hours=3),
                output_field=DateTimeField()
            )
        )
        .annotate(day=TruncDate('shifted_time'))
        .values('day')
        .annotate(total_qty=Sum('quantity'))
        .order_by('day')
    )

    customers_count = Invoice.objects.values('phone').distinct().count()

    context = {
        'top_customers': top_customers,
        'daily_sales': daily_sales,
        'total_revenue': summary['total_revenue'] or 0,
        'days_count': len(daily_sales),
        'customers_count': customers_count,
        'daily_item_volume': daily_item_volume,
    }
    return render(request, 'factors_data_dashboard.html', context)







from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import Invoice
from .utils import check_personel_noght_order, extract_payment_methods, get_persian_date_string, jalali_date_time_to_gregorian, normalize_iranian_mobile, read_excel



class ReceiveInvoice(APIView):
    authentication_classes = []
    permission_classes = []
    print('req reciev'*20)
    def post(self, request):
        data = request.data
        if request.headers.get("X-API-KEY") != "SECRET123":
            return Response({"error": "unauthorized"}, status=403)
    

        print(f"DEBUG: Invoice Number Type: {type(data['invoice_number'])}")
        print(f"DEBUG: Invoice Number Value: '{data['invoice_number']}'")

        
        date_time = jalali_date_time_to_gregorian(data['date'],data['time'])
        exists = Invoice.objects.filter(invoice_number=data["invoice_number"]).exists()
        print(f"DEBUG: Record exists in DB? {exists}")
        # استفاده از update_or_create بر
        
        
        try:
            invoice_num = int(data["invoice_number"])
            # چون فیلد شما CharField است، دوباره آن را به استرینگ تمیز تبدیل می‌کنیم
            # تا صفرهای اضافه حذف شوند (مثلاً '00123' تبدیل به '123' می‌شود)
            invoice_num_str = str(invoice_num) 
        except (ValueError, TypeError):
            # اگر شماره فاکتور اصلاً عدد نیست، همان مقدار اصلی را بفرستید یا خطا دهید
            invoice_num_str = str(data["invoice_number"]).strip()

        
        # Invoice.objects.filter(invoice_number=data["invoice_number"]).delete()

        year_shamsi = jdatetime.datetime.now().year
        unique_invoice_number = f'{year_shamsi}_{data["invoice_number"]}'
        invoice, created_invoice = Invoice.objects.update_or_create(
            unique_invoice_number=unique_invoice_number,
            defaults={
                "invoice_number":data["invoice_number"],
                "name": data["name"],
                "nahveh": data["nahveh"],
                "phone": data["phone"],
                "created_at": date_time,
                "total_price": data["total_price"],
                "discount" : data['takhfif'],
                "peyk" : data['peyk'],
                # "anaam" : data['anaam'],

                "moshtarak" : data['moshtarak'],
                "serv" : data['serv'],
                "pnum" : data['pnum'],
                "shomare_pos" : data['shomareh_pos'],
                "mablagh_pos" : data['mablagh_pos'],
                "hazine_peyk" : data['hazineh_peyk'],
                "naghdi" : data['naghdi'],
                "nonaghdi" : data['nonaghdi'],
                "mandeh" : data['mandeh'],

            }
        )
        print('created : ',created_invoice,'Invoice :',data["invoice_number"])
        # اگر فاکتور قبلاً وجود داشته است، آیتم‌های قدیمی آن را پاک می‌کنیم تا آیتم‌های جدید جایگزین شوند
      
        InvoiceItem.objects.filter(invoice=invoice).delete()
            
        for item in data["items"]:
            InvoiceItem.objects.create(
                invoice=invoice,
                food_name=item["food"],
                price=item["price"],
                quantity=item["quantity"],
                total=item["price"] * item["quantity"]
            )
            
            
        sms_log, created = SMSLog.objects.get_or_create(invoice_number=data["invoice_number"])
        


        now = timezone.now()
        time_condition=True
        if now - date_time >= timedelta(hours=3):
            time_condition = False
            
  
        print('time_condition :',time_condition)


        if time_condition:
            if created_invoice:
                if not sms_log.is_sent:
                    sms_log.is_sent = True
                    sms_log.save()
                    print(os.getenv('SMS_NIGHT_ORDER'))
                    if os.getenv('SMS_NIGHT_ORDER'):
                        moshtarak =  str(data['moshtarak'])
                        print('moshtarak',moshtarak)
                        profile_moshtarak = Profile.objects.filter(code_vaset=moshtarak).first()
                        if profile_moshtarak is None:
                            pass
                        else:
                            ret , phones =  normalize_iranian_mobile(profile_moshtarak.phone)
                            
                            ret , phones = check_personel_noght_order(ret,phones,data)
                            
                            print('phones',phones)
                            
                            
                            if ret:
                                current_time = timezone.now()
                                selected_date = current_time.date()
                                persian_date_str = get_persian_date_string(selected_date)
                            
                                sms_template = SMS_Template.objects.filter(name =SMSServiceTemplate_Enum.PERSONEL_NIGHT_ORDER )
                                if sms_template.exists():
                                    name = profile_moshtarak.full_name
                                    sms_template = sms_template.first()
                                    for phone in phones:
                                        if phone:
                                            pass
                                            ret = send_sms(sms_template,phone_number=phone,vars={OTPVar_Enum.NAME:name,OTPVar_Enum.DATE:persian_date_str,OTPVar_Enum.ORDERID:data["invoice_number"],OTPVar_Enum.PRICE:data["total_price"],})
                                            print('SMS Send to : ',phone, 'ret:',ret)
                                        else:
                                            print('Phone not Exist for send')


                


        return Response({"status": "ok"}, status=status.HTTP_201_CREATED)
    



from django.db.models import Max
from collections import defaultdict
from datetime import datetime
import re
from django.db.models import Max
from collections import defaultdict
from datetime import datetime, timedelta, time
import re

def calc_nahve_pardakh(request):
    date_str = request.GET.get('date')

    # 🔹 اگر تاریخ نفرستاده بود → آخرین تاریخ دیتابیس
    if date_str:
        selected_date = datetime.strptime(date_str, '%Y/%m/%d').date()
    else:
        last_date = Invoice.objects.aggregate(
            max_date=Max('created_at')
        )['max_date']

        if not last_date:
            return render(request, 'utils/nahve.html', {})

        selected_date = last_date.date()

    # 🔹 بازه روز کاری: 3 صبح تا 3 صبح روز بعد
    start_datetime = datetime.combine(selected_date, time(3, 0))
    end_datetime = start_datetime + timedelta(days=1)

    invoices = Invoice.objects.filter(
        created_at__gte=start_datetime,
        created_at__lt=end_datetime
    ).prefetch_related("items")


    totals = defaultdict(int)

    total_items = 0

    for invoice in invoices:
        print(invoice.name)
        if 'کیوسک۱' in invoice.name:
            totals['کیوسک۱'] += invoice.total_price# -invoice.discount
        elif 'کیوسک۲' in invoice.name:
            totals['کیوسک۲'] += invoice.total_price# -invoice.discount
        elif 'کیوسک۳' in invoice.name:
            totals['کیوسک۳'] += invoice.total_price# -invoice.discount


        elif 'اسنپ' in invoice.name:
            totals['اسنپ'] += invoice.total_price# -invoice.discount


        else:
            methods = extract_payment_methods(invoice.nahveh)
            for method in methods:
                totals[method] += invoice.total_price #- invoice.discount
        total_items+= len(invoice.items.all())

    context = {
        'selected_date': selected_date,
        'labels': list(totals.keys()),
        'values': list(totals.values()),
        'table_data': totals.items(),
        'total_items':total_items
    }

    return render(request, 'utils/nahve.html', context)






from django.shortcuts import render
from django.http import HttpResponse, JsonResponse
from datetime import datetime
import pandas as pd

from .models import Invoice, InvoiceItem
from .utils import get_date_range  # or paste function directly


def invoice_report(request):
    # default = today
    date_str = request.GET.get('date')

    if date_str:
        selected_date = datetime.strptime(date_str, "%Y-%m-%d").date()
    else:
        selected_date = timezone.now().date()

    start, end = get_date_range(selected_date)

    invoices = Invoice.objects.filter(
        created_at__range=(start, end)
    ).prefetch_related("items")

    context = {
        "invoices": invoices,
        "selected_date": selected_date
    }
    return render(request, "invoice_report.html", context)

from dateutil import parser


def download_invoice_excel(request):
    date_str = request.GET.get('date')
    if date_str:
        try:
            selected_date = parser.parse(date_str).date()
        except (ValueError, TypeError):
            selected_date = datetime.today().date()
    else:
        selected_date = datetime.today().date()


    # selected_date = datetime.strptime(date_str, "%Y-%m-%d").date() if date_str else timezone.localdate()

    start, end = get_date_range(selected_date)

    rows = []

    items = InvoiceItem.objects.filter(
        invoice__created_at__range=(start, end)
    ).select_related("invoice")

    for item in items:
        rows.append({
            "Invoice Number": item.invoice.invoice_number,
            "Name": item.invoice.name,
            "Phone": item.invoice.phone,
            "Food": item.food_name,
            "Price": item.price,
            "Quantity": item.quantity,
            "Total": item.total,
            "Date": item.invoice.created_at
        })

    df = pd.DataFrame(rows)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="invoice_detail.xlsx"'

    df.to_excel(response, index=False)
    return response




def download_invoice_summary_excel(request):
    date_str = request.GET.get('date')
    selected_date = datetime.strptime(date_str, "%Y-%m-%d").date() if date_str else timezone.localdate()

    start, end = get_date_range(selected_date)

    items = InvoiceItem.objects.filter(
        invoice__created_at__range=(start, end)
    )

    rows = [{
        "Food": i.food_name,
        "Quantity": i.quantity,
        "Total": i.total
    } for i in items]

    df = pd.DataFrame(rows)

    summary = df.groupby("Food", as_index=False).sum()

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="invoice_summary.xlsx"'

    summary.to_excel(response, index=False)
    return response



def invoice_detail_api(request, invoice_number):
    invoice = get_object_or_404(
        Invoice.objects.prefetch_related("items"),
        invoice_number=invoice_number
    )

    data = {
        "invoice_number": invoice.invoice_number,
        "name": invoice.name,
        "phone": invoice.phone,
        "nahveh": invoice.nahveh,
        "created_at": invoice.created_at.strftime("%Y-%m-%d %H:%M"),
        "total_price": invoice.total_price,
        "discount": invoice.discount,
        "items": [
            {
                "food_name": item.food_name,
                "price": item.price,
                "quantity": item.quantity,
                "total": item.total,
            }
            for item in invoice.items.all()
        ]
    }

    return JsonResponse(data)







def vaset_convert_peyk_details(peyk_id:int):

    data =  read_excel(excel_name='peyk.xlsx')
    # مثال: چاپ سلول A1


    select1 = None
    select2 = None
    select3 = None

    for d in data:
        try:
            if 'متفرقه' in str(d[0]):
                select1 = int(d[2])
                select2 = int(d[3])
                select3 = int(d[4])
            if int(d[1])== int(peyk_id):
                if d[2] is not None and d[3] is not None and d[4] is not None:
                    return int(d[2]),int(d[3]),int(d[4])
        except :
            pass
    
    return select1,select2,select3


from io import BytesIO
from pathlib import Path
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from django.http import HttpResponse
from django.utils import timezone
from .folder_utils.foodsoft_lookup import get_kname_by_kcod
from .folder_utils.sepidar_lookup import get_code_by_name

MOSHTARAK_DEFAULT_CODE = 10012
FOODSOFT_SNAPP_CODE = 1
SEPIDAR_SNAPP_CODE = 10059

CODE_MOIN_POS_MOTOFAREGHE = 121304
MOSTARAK_FANTEZI = 20142
POS_FANTEZI = 11



def sepidar_download_excel(request):
    SERVER = check_server()

    date_str = request.GET.get('date')
    is_total = request.GET.get('total', '').lower() == 'true'
    if date_str:
        try:
            selected_date = parser.parse(date_str).date()
        except (ValueError, TypeError):
            selected_date = datetime.today().date()
    else:
        selected_date = datetime.today().date()


    # selected_date = datetime.strptime(date_str, "%Y-%m-%d").date() if date_str else timezone.localdate()

    start, end = get_date_range(selected_date)

    invoices = (
        Invoice.objects
        .filter(created_at__range=(start, end))
        .prefetch_related("items")
    )



    # -----------------------------
    # 1) Build rows (one row per invoice item)
    # -----------------------------
    rows = []
    final_factors_count = 0
    invoice_total_price = 0
    snapp_total_price = 0
    motofareghe_total_price = 0

    errors = []
    error_factors = []  # List to track which invoices had errors


 
    codes = (Profile.objects
            .exclude(code_vaset__isnull=True)
            .exclude(code_vaset__exact='')
            .values_list('code_vaset', flat=True)
            .distinct()
            .order_by('code_vaset'))
    codes = list(codes)




    for inv in invoices:

        tasvie_model = 1

        it_discount_flag = False

        if float(inv.mandeh)>0:
            if float(inv.mandeh)== float(inv.total_price)+float(inv.hazine_peyk)-float(inv.discount) :
                tasvie_model = 2
            else:
                tasvie_model = 3

        if float(inv.hazine_peyk)>0 and int(inv.moshtarak)!=FOODSOFT_SNAPP_CODE:
            tasvie_model = 1

        cancel=False
        empty_factor = True

        for it in inv.items.all():
            empty_factor = False
            if it.food_name == '':
                pass
            name = get_kname_by_kcod(it.food_name)
            code = get_code_by_name(name=name)
            try:
                _id = int(it.food_name)
            except:
                _id=0
            try:
                if int(code) == 2500013 or _id ==65 :
                    cancel = True
                else:
                    break
            except:
                error_factors.append(f'Error in Check emoty factor :{inv.invoice_number} ' )  # Add invoice number to error list

                print("error in empty factor check")
        if cancel:
            inv.discount = inv.total_price
            # continue
        
        if empty_factor:
            continue

        if inv.created_at.hour<4:
            date = (inv.created_at - timezone.timedelta(days=1)).replace(hour=23, minute=59, second=59, microsecond=0)
        else:
            date = inv.created_at
        date_time = format_jalali_datetime(date)


        for it in inv.items.all():

            empty_factor=False
            name = get_kname_by_kcod(it.food_name)
            if name is None or name=='':
                error_factors.append(f'{str(inv.invoice_number)} , Food in fodsoft with name {name} not exist in Code'  )  # Add invoice number to error list

                pass
            code = get_code_by_name(name=name)
            if code is None:
                error_text = f'Error food soft : Empty code {code} , it.food_name : {it.food_name} , name : {name}'
                if not SERVER:
                    print(error_text)
                errors.append(error_text)
                invoice_has_error = True  # Mark invoice as having error
                error_factors.append(f'{str(inv.invoice_number)} , {error_text}'  )  # Add invoice number to error list


                

            try:
                int(inv.moshtarak)
            except:
                inv.moshtarak = 0
            if int(inv.moshtarak) == FOODSOFT_SNAPP_CODE:
                moshtarak = SEPIDAR_SNAPP_CODE  ### SNAPP DEFAULT CODE
            elif 10000<int(inv.moshtarak)<=20000:
                moshtarak = inv.moshtarak
            # elif int(inv.shomare_pos) == POS_FANTEZI :
            #     if inv.created_at >= datetime(2026, 4, 21):
            #         moshtarak = MOSTARAK_FANTEZI
            #     else:
                    # moshtarak = MOSHTARAK_DEFAULT_CODE

            else:
                moshtarak = MOSHTARAK_DEFAULT_CODE

            it.discount=0

      

            if int(it.total)>=int(inv.discount) and not it_discount_flag:
                it.discount = inv.discount
                it_discount_flag = True
            
            data = {
                'نوع قلم' : 'InvoiceItem',
                "فاكتور شماره": inv.invoice_number,
                "فاكتور نام مشتري": inv.name,
                'فاكتور كد مشتري': moshtarak,
                'فاكتور تخفيف': inv.discount,
                'فاكتور كد نوع فروش': 1,
                # "phone": inv.phone,
                "قلم فاكتور كد": code,
                "فاكتور كل": inv.total_price,
                "قلم فاكتور في": it.price,
                "قلم فاكتور واحد اصلي": it.quantity,
                'قلم فاكتور تخفيف مشتري':it.discount,
                'قلم فاكتور كد انبار':5,
                "قلم فاكتور كل": it.total,
                "فاكتور تاريخ": date_time,
                'فاكتور ارز1':'ريال',
                'فاكتور نرخ ارز':1,
                'فاكتور نوع تسويه':tasvie_model,
                'فاكتور محل  تحويل':'آدرس مشتري',
                'اطلاعات مبالغ دريافتي فاكتور مبلغ حواله' : float(inv.mablagh_pos),
                'اطلاعات مبالغ دريافتي فاكتور مبلغ نقد' : float(inv.naghdi),

            }

            rows.append(data)



        final_factors_count+=1
        invoice_total_price += float(inv.total_price)

        if moshtarak==SEPIDAR_SNAPP_CODE:
            snapp_total_price += float(inv.total_price)
        if moshtarak == MOSHTARAK_DEFAULT_CODE:
            motofareghe_total_price += float(inv.total_price)


        if float(inv.peyk>0) or float(inv.hazine_peyk)>0 :
            if inv.peyk>6:
                pass
            peyk_code,peyk_tafzil,peyk_vaset = vaset_convert_peyk_details(inv.peyk)
            if peyk_tafzil==157:
                pass

            peyk_calc_code = 9200001

            data = {
                    'نوع قلم' : 'InvoiceItem',
                    "فاكتور شماره": inv.invoice_number,
                    "فاكتور نام مشتري": inv.name,
                    'فاكتور كد مشتري': moshtarak,
                    'فاكتور تخفيف': inv.discount,
                    'فاكتور كد نوع فروش': 1,
                    # "phone": inv.phone,
                    "قلم فاكتور كد": peyk_calc_code,
                    "فاكتور كل": inv.total_price ,
                    "قلم فاكتور في": int(float(inv.hazine_peyk)),
                    "قلم فاكتور واحد اصلي": 1,
                    'قلم فاكتور تخفيف مشتري':0,
                    'قلم فاكتور كد انبار':'',
                    "قلم فاكتور كل": int(float(inv.hazine_peyk)),
                    "فاكتور تاريخ": date_time,
                    'فاكتور ارز1':'ريال',
                    'فاكتور نرخ ارز':1,
                    'فاكتور نوع تسويه':tasvie_model,
                    'فاكتور محل  تحويل':'آدرس مشتري',
                    'اطلاعات مبالغ دريافتي فاكتور مبلغ حواله' : float(inv.mablagh_pos),
                    'اطلاعات مبالغ دريافتي فاكتور مبلغ نقد' : float(inv.naghdi),

                }

            rows.append(data)

            # if int(inv.peyk) == 4: ## Javad Zamani
                # hazine_peyk = int(float(inv.hazine_peyk))*1
            # else:
            hazine_peyk = int(float(inv.hazine_peyk))*0.9

            rows.append({
                'نوع قلم' : 'InvoiceBroker',
                "فاكتور شماره": inv.invoice_number,
                'فاكتور كد نوع فروش': 1,
                "فاكتور نام مشتري": inv.name,
                'فاكتور كد مشتري': moshtarak,
                # "phone": inv.phone,
                'واسط مبلغ پورسانت':hazine_peyk,
                'واسط تفصيلي واسط':peyk_code,
                'واسط كد واسط':peyk_tafzil,
                'واسط واسط':peyk_vaset,
                "فاكتور تاريخ": date_time,
                'فاكتور ارز1':'ريال',
                'فاكتور نرخ ارز':1,
                'فاكتور نوع تسويه':tasvie_model,
                'فاكتور محل  تحويل':'آدرس مشتري',

            })

    # -----------------------------
    # 2) Template + mapping
    # -----------------------------
    if is_total:
        rows = sepidar_factor_total(rows)

    print('final_factors_count : ',final_factors_count)
    print('invoice_total_price : ',invoice_total_price)
    print('motofareghe_total_price : ',motofareghe_total_price)
    print('snapp_total_price : ',snapp_total_price)

    # Load once at import time
    SERVER = check_server()
    if SERVER:
        template_path = r"/home/seketal1/Seketala_Kitchen_Flow/cache/sepidar_template.xlsx"  # must be .xlsx
    else:
        template_path = r'cache\sepidar_template.xlsx'
    wb = load_workbook(template_path)
    ws = wb[wb.sheetnames[0]]  # or wb["Sheet1"]

    START_ROW =2 # where the first data row begins in your template

    # Only these columns will be filled; everything else stays unchanged
    # مثال: اگر نمیخوای ستون C پر بشه، اصلا اینجا قرارش نده
    COL_MAP = {
        'نوع قلم' : 'A',
        "فاكتور شماره": 'B',
        "فاكتور تاريخ": 'C',
        'فاكتور كد مشتري': 'D',
        'فاكتور كد نوع فروش': 'E',
        "قلم فاكتور كد": 'F',
        'قلم فاكتور كد انبار':'G',
        "قلم فاكتور واحد اصلي": 'I',
        'قلم فاكتور تخفيف مشتري':'T',

        "قلم فاكتور في": 'K',
        "قلم فاكتور كل": 'L',
        "فاكتور نام مشتري": 'R',
        'فاكتور محل  تحويل':'S',

        'فاكتور تخفيف': 'Y',
        # "phone": inv.phone,
        'واسط مبلغ پورسانت':'Z',
        'فاكتور ارز1':'V',
        'فاكتور نرخ ارز':'W',
        'فاكتور نوع تسويه':'X',

        'واسط كد واسط':'AA',

        'واسط تفصيلي واسط':'AB',
        'واسط واسط':'AD',
        "فاكتور كل":'AC',
        'اطلاعات مبالغ دريافتي فاكتور مبلغ حواله' : 'AE',
        'اطلاعات مبالغ دريافتي فاكتور مبلغ نقد' : 'AF',
    }

    # Precompute numeric column indexes (faster)
    col_idx_map = {k: column_index_from_string(v) for k, v in COL_MAP.items()}

    # -----------------------------
    # 3) Write only mapped columns
    # -----------------------------
    for i, record in enumerate(rows):
        excel_row = START_ROW + i
        for field, col_idx in col_idx_map.items():
            ws.cell(row=excel_row, column=col_idx, value=record.get(field))

    # -----------------------------
    # 4) Return as download
    # -----------------------------
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    output =  convert_xlsx2xls(output=output)



    j_date = jdatetime.date.fromgregorian(date=selected_date)
    filename = f"sepidar_{j_date.strftime('%Y-%m-%d')}.xls"  # تغییر پسوند به .xls
    

    resp = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.ms-excel",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    
    # Add errors to response headers (only if there are errors)
    if error_factors:
        # Remove duplicates and limit to first 50 to avoid header size issues
        unique_errors = list(set(error_factors))[:50]
        import json
        resp["X-Error-Factors"] = json.dumps(unique_errors)
    
    return resp



def sepidar_factor_total(rows):
    from collections import defaultdict

    MERGE_CUSTOMER_CODES = [
        SEPIDAR_SNAPP_CODE,
        MOSHTARAK_DEFAULT_CODE,
]

    broker_items = defaultdict(
        lambda: defaultdict(
            lambda: {
                "commission": 0
            }
        )
    )

    new_rows = []

    merged_data = {}

    for customer_code in MERGE_CUSTOMER_CODES:
        merged_data[customer_code] = {
            "items": defaultdict(lambda: {
                "qty": 0,
                "total": 0,
                "takhfif": 0,
            }),
            "discount": 0,
            "pos": 0,
            "cash": 0,
            "invoice": 0,
            "peyk": 0,
            "first_row": None,
        }



    for row in rows:

        customer_code = int(row.get('فاكتور كد مشتري', 0))

        if customer_code not in MERGE_CUSTOMER_CODES:
            new_rows.append(row)
            continue

        data = merged_data[customer_code]

        if data["first_row"] is None:
            data["first_row"] = row.copy()

        if row.get('نوع قلم') == 'InvoiceItem':

            product_code = row['قلم فاكتور كد']

            if product_code is None:
                continue

            qty = float(row.get('قلم فاكتور واحد اصلي', 0))
            total = float(row.get('قلم فاكتور كل', 0))
            takhfif = float(row.get('قلم فاكتور تخفيف مشتري', 0))

            data["items"][product_code]["qty"] += qty
            data["items"][product_code]["total"] += total
            data["items"][product_code]["takhfif"] += takhfif

            data["discount"] += float(
                row.get('قلم فاكتور تخفيف مشتري', 0)
            )

            data["pos"] += float(
                row.get('اطلاعات مبالغ دريافتي فاكتور مبلغ حواله', 0)
            )

            data["cash"] += float(
                row.get('اطلاعات مبالغ دريافتي فاكتور مبلغ نقد', 0)
            )

            data["invoice"] += float(
                row.get('فاكتور كل', 0)
            )

            if product_code == 9200001:
                data["peyk"] += total


        elif row.get('نوع قلم') == 'InvoiceBroker':


            peyk_code = row.get('واسط تفصيلي واسط')
            peyk_tafzil = row.get('واسط كد واسط')
            peyk_vaset = row.get('واسط واسط')

            key = (
                peyk_code,
                peyk_tafzil,
                peyk_vaset
            )

  

            if int(customer_code) == 10059:
                pass

            # print('customer_code',customer_code,'key',key)

            broker_items[customer_code][key]["commission"] += float(
                row.get('واسط مبلغ پورسانت', 0) or 0
            )

    # ساخت فاکتور تجمیعی هر مشتری
    for customer_code in MERGE_CUSTOMER_CODES:

        data = merged_data[customer_code]

        if data["first_row"] is None:
            continue

        customer_name = (
            'اسنپ تجمیعی'
            if customer_code == SEPIDAR_SNAPP_CODE
            else f'تجمیعی {customer_code}'
        )

        for product_code, item  in data["items"].items():

            qty = item["qty"]
            total = item["total"]
            takhfif = item["takhfif"]
            if int(product_code) == 9200001:
                pass
            fee = np.round(total / qty if qty else 0)

            new_rows.append({
                'نوع قلم': 'InvoiceItem',

                'فاكتور شماره': data["first_row"]['فاكتور شماره'],
                'فاكتور نام مشتري': customer_name,
                'فاكتور كد مشتري': customer_code,

                'فاكتور تخفيف': data["discount"],
                'فاكتور كد نوع فروش': 1,

                'قلم فاكتور كد': product_code,
                'قلم فاكتور في': round(fee),
                'قلم فاكتور واحد اصلي': qty,
                'قلم فاكتور تخفيف مشتري': takhfif,
                'قلم فاكتور كد انبار': 5,
                'قلم فاكتور كل': total,

                'فاكتور كل': data["invoice"],

                'اطلاعات مبالغ دريافتي فاكتور مبلغ حواله': data["pos"],
                'اطلاعات مبالغ دريافتي فاكتور مبلغ نقد': data["cash"],

                'فاكتور ارز1': 'ريال',
                'فاكتور نرخ ارز': 1,
                'فاكتور نوع تسويه': 1,
                'فاكتور محل  تحويل': 'آدرس مشتري',

                'فاكتور تاريخ': data["first_row"]['فاكتور تاريخ']
            })

        # if customer_code==MOSHTARAK_DEFAULT_CODE:

        # for (
        #         peyk_code,
        #         peyk_tafzil,
        #         peyk_vaset
        #     ), item in broker_items.items():
        if customer_code in broker_items and broker_items[customer_code]:

            for (peyk_code, peyk_tafzil, peyk_vaset), item in broker_items[customer_code].items():
                # print('new proker', (peyk_code, peyk_tafzil, peyk_vaset))
                hazine_peyk = item["commission"]

                new_rows.append({
                    'نوع قلم': 'InvoiceBroker',

                    'فاكتور شماره': data["first_row"]['فاكتور شماره'],
                    'فاكتور نام مشتري': customer_name,
                    'فاكتور كد نوع فروش': 1,
                    'فاكتور كد مشتري': customer_code,

                    'واسط تفصيلي واسط': peyk_code,
                    'واسط كد واسط': peyk_tafzil,
                    'واسط واسط': peyk_vaset,

                    'واسط مبلغ پورسانت': hazine_peyk,

                    'فاكتور تاريخ': data["first_row"]['فاكتور تاريخ'],

                    'فاكتور ارز1': 'ريال',
                    'فاكتور نرخ ارز': 1,
                    'فاكتور نوع تسويه': 1,
                    'فاكتور محل  تحويل': 'آدرس مشتري',
                })

    return new_rows

def convert_xlsx2xls(output):

    wb_xlsx = load_workbook(output)
    
    # ایجاد فایل XLS جدید با xlwt
    wb_xls = xlwt.Workbook()
    
    # کپی تمام صفحات به فرمت XLS
    for sheet_name in wb_xlsx.sheetnames:
        ws_xlsx = wb_xlsx[sheet_name]
        ws_xls = wb_xls.add_sheet(sheet_name)
        
        # کپی داده‌ها از XLSX به XLS
        for row_idx, row in enumerate(ws_xlsx.iter_rows()):
            for col_idx, cell in enumerate(row):
                ws_xls.write(row_idx, col_idx, cell.value)
    
    # ذخیره فایل XLS در BytesIO جدید
    output_xls = BytesIO()
    wb_xls.save(output_xls)
    output_xls.seek(0)

    return output_xls


def bank_vaset(pos_id:int):

    data =  read_excel(excel_name='Bank_vaset.xlsx')


    select1 = None
    select2 = None

    for d in data:
        try:
            if int(d[0])== int(pos_id):
                if d[1] is not None and d[2] is not None :
                    return int(d[1]),str(d[2])
        except :
            pass
    
    return select1,select2









def tasvieh_sepidar_download_excel(request):
    date_str = request.GET.get('date')
    is_total = request.GET.get('total', '').lower() == 'true'

    if date_str:
        try:
            selected_date = parser.parse(date_str).date()
        except (ValueError, TypeError):
            selected_date = datetime.today().date()
    else:
        selected_date = datetime.today().date()


    # selected_date = datetime.strptime(date_str, "%Y-%m-%d").date() if date_str else timezone.localdate()

    jalali = jdatetime.datetime.fromgregorian(datetime=selected_date)
    jalali_date = jalali.strftime("%Y-%m-%d")
    # Various formats
    # print(jalali.strftime("%Y-%m-%d"))           # 1405-04-15


    start, end = get_date_range(selected_date)

    invoices = (
        Invoice.objects
        .filter(created_at__range=(start, end))
        .prefetch_related("items")
    )


    error_factors = []



    # -----------------------------
    # 1) Build rows (one row per invoice item)
    # -----------------------------
    rows = []
    SKIPPED_SHOMARE_POS = [11,13,14]
    for inv in invoices:

        tafsil_bank,havale_bank = None , None

        if int(inv.moshtarak) ==1:
            continue
        if float(inv.mandeh)>0:
            continue

        today_payment = False

        shomare_pos = int(inv.shomare_pos)

        if shomare_pos in SKIPPED_SHOMARE_POS:
            continue




        try:
            if float(inv.nonaghdi)>0:
                inv.mablagh_pos = inv.nonaghdi
                if '[واريز به کارت ملي 2' in  inv.nahveh or '[واريز به حساب ملي 2'  in  inv.nahveh:
                    tafsil_bank,havale_bank = bank_vaset(int(31))
                    today_payment = True

                elif '[واريز به کارت ملي 1' in  inv.nahveh or  '[واريز به حساب ملي 1' in  inv.nahveh:
                    tafsil_bank,havale_bank = bank_vaset(int(32))    
                    today_payment = True

                elif '[واريز به کارت ملي 9' in  inv.nahveh or  '[واريز به حساب ملي 9' in  inv.nahveh:
                    tafsil_bank,havale_bank = bank_vaset(int(33))    
                    today_payment = True


                else:
                    error_factors.append(f'خطا در فاکتور :  {inv.invoice_number} , {inv.nahveh} در فاکتور ثبت شده')
                    print('Error New type Detected')
                    continue
        except:
            pass

        if float(inv.naghdi)>0 :
            naghdi = int(float(inv.naghdi))
            if float(inv.mablagh_pos)<=0 :
                today_payment = True

        else:
            naghdi = 0 
            
        if float(inv.mablagh_pos)>0:
            mablagh_pos = int(float(inv.mablagh_pos))
            if tafsil_bank is None and havale_bank is None:
                tafsil_bank,havale_bank = bank_vaset(int(inv.shomare_pos))
            type= 'ReceiptDraft'
            sandogh = ''

            if tafsil_bank is None or havale_bank is None:
                error_factors.append(f'{inv.invoice_number} مبلغ : {inv.mablagh_pos} , شماره پوز : {inv.shomare_pos} , نحوه تسویه : {inv.nahveh}')
                continue

        else:
            mablagh_pos = ''
            type = ''
            sandogh = 'صندوق مركزي'

        if today_payment:
            payment_date = format_jalali_datetime(inv.created_at)
        else:
            payment_date = havale_format_jalali_datetime(inv.created_at)
  
    
        if 10000<int(inv.moshtarak)<=20000:
            taraf_moghabel = inv.moshtarak
        else:
            # 'متفرقه/فروش'
            taraf_moghabel = MOSHTARAK_DEFAULT_CODE




        rows.append({
            'نوع قلم' : type,
            'رسيد دريافت نوع دريافت':1,
            'رسيد دريافت کد':taraf_moghabel,
            'رسيد دريافت شماره':inv.invoice_number,
            'رسيد دريافت تاريخ':payment_date,
            'رسيد دريافت كد معين':121201,
            'رسيد دريافت صندوق':sandogh,
            'رسيد دريافت مبلغ نقد':naghdi,
            'رسيد دريافت شرح':f'فاكتور شماره {inv.invoice_number}',
            'رسيد دريافت مبلغ دريافت':inv.total_price-inv.discount,
            'حواله شماره':inv.invoice_number,
            'حواله تاريخ':payment_date,
            'حواله مبلغ':mablagh_pos,
            'حواله شرح':f'فاكتور شماره {inv.invoice_number}',
            'حواله تفصيل حساب بانكي':tafsil_bank,
            'حواله حساب بانکی':havale_bank,
            'رسید دریافت تخفیف' : 0,
            'رسید دریافت استقراری' : 'False'

        })






    # -----------------------------
    # 2) Template + mapping
    # -----------------------------

    

    if is_total:
        rows = sepidar_resid_total(data=rows,selected_date=jalali_date)

    # Load once at import time
    SERVER = check_server()
    if SERVER:
        template_path = r"/home/seketal1/Seketala_Kitchen_Flow/cache/resid_sepidar_template.xlsx"  # must be .xlsx
    else:
        template_path = r'cache\resid_sepidar_template.xlsx'
    wb = load_workbook(template_path)
    ws = wb[wb.sheetnames[0]]  # or wb["Sheet1"]

    START_ROW =2 # where the first data row begins in your template

    # Only these columns will be filled; everything else stays unchanged
    # مثال: اگر نمیخوای ستون C پر بشه، اصلا اینجا قرارش نده
    COL_MAP = {
        'نوع قلم' : 'A',
        'رسيد دريافت نوع دريافت':'B',
        'رسيد دريافت کد':'C',
        'رسيد دريافت شماره':'D',
        'رسيد دريافت تاريخ':'E',
        'رسيد دريافت كد معين':'F',
        'رسيد دريافت صندوق':'G',
        'رسيد دريافت مبلغ نقد':'H',
        'رسيد دريافت شرح':'I',
        'رسيد دريافت مبلغ دريافت':'AA',
        'حواله شماره':'AB',
        'حواله تاريخ':'AC',
        'حواله مبلغ':'AD',
        'حواله شرح':'AF',
        'حواله تفصيل حساب بانكي':'AH',
        'رسید دریافت تخفیف' : 'AL',
        'رسید دریافت استقراری' : 'AM',    
        'حواله حساب بانکی':'AE',
        # "date": "M",   # if you don't want it, comment/remove it
    }

    # Precompute numeric column indexes (faster)
    col_idx_map = {k: column_index_from_string(v) for k, v in COL_MAP.items()}

    # -----------------------------
    # 3) Write only mapped columns
    # -----------------------------
    for i, record in enumerate(rows):
        excel_row = START_ROW + i
        for field, col_idx in col_idx_map.items():
            ws.cell(row=excel_row, column=col_idx, value=record.get(field))

    # -----------------------------
    # 4) Return as download
    # -----------------------------
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    output = convert_xlsx2xls(output=output)

    j_date = jdatetime.date.fromgregorian(date=selected_date)

    filename = f"resid_sepidar_{j_date.strftime('%Y-%m-%d')}.xls"
    resp = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.ms-excel",  # مقدار جدید برای XLS
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    resp["X-Error-Factors"] = json.dumps(error_factors)
    return resp

def sepidar_resid_total(data,selected_date):
    """
    پردازش داده‌ها در حالت total:
    - جمع‌آوری مبالغ نقدی و کارت به کارت به تفکیک تاریخ (فقط روز) و شماره بانک
    - بازگشت ساختار جدید با سطرهای تجمعی به تفکیک تاریخ
    """
    from collections import defaultdict
    import re
    
    def extract_date_only(date_string):
        """
        استخراج فقط تاریخ (بدون ساعت) از رشته تاریخ
        """
        if not date_string:
            return ''
        
        # اگر تاریخ به صورت 1405/03/14 باشد
        if re.match(r'^\d{4}/\d{2}/\d{2}$', date_string):
            return date_string
        
        # اگر تاریخ شامل ساعت باشد: 1405/03/14 19:17:00 ب.ظ
        match = re.match(r'^(\d{4}/\d{2}/\d{2})', date_string)
        if match:
            return match.group(1)
        
        # اگر تاریخ به فرمت دیگری باشد
        return date_string
    
    # دیکشنری برای ذخیره مبالغ تجمعی با کلید ترکیبی (تاریخ, شماره بانک)
    total_data = defaultdict(lambda: {
        'naghdi_total': 0,
        'kart_be_kart_total': 0,
        'invoices': [],
        'first_invoice': None
    })



    bankd_detail = {}
    
    other_rows = []

    for row in data:
        # استخراج تاریخ و حذف ساعت
        if row['رسيد دريافت کد']!= MOSHTARAK_DEFAULT_CODE:
            other_rows.append(row)
            continue
        
        
        payment_date = row.get('رسيد دريافت تاريخ', '')
        if payment_date:
            payment_date = extract_date_only(payment_date)
        else:
            payment_date = row.get('حواله تاريخ', '')
            if payment_date:
                payment_date = extract_date_only(payment_date)
        
        # استخراج شماره بانک
        bank_code = row.get('حواله تفصيل حساب بانكي', '')
        # if not bank_code:
        bank_name = row.get('حواله حساب بانکی', '')
        bankd_detail.update({bank_code:bank_name})
        if not bank_code:
            continue
        
        # ایجاد کلید ترکیبی (تاریخ, شماره بانک)
        key = (payment_date, bank_code)
        
        # اضافه کردن مبلغ نقدی
        naghdi = row.get('رسيد دريافت مبلغ نقد', 0)
        if isinstance(naghdi, str) and naghdi:
            try:
                naghdi = int(float(naghdi))
            except (ValueError, TypeError):
                naghdi = 0
        
        # اضافه کردن مبلغ کارت به کارت (حواله)
        kart_be_kart = row.get('حواله مبلغ', 0)
        if isinstance(kart_be_kart, str) and kart_be_kart:
            try:
                kart_be_kart = int(float(kart_be_kart))
            except (ValueError, TypeError):
                kart_be_kart = 0
        
        # جمع‌آوری در دیکشنری
        total_data[key]['naghdi_total'] += naghdi
        total_data[key]['kart_be_kart_total'] += kart_be_kart
        total_data[key]['invoices'].append(row)
        if total_data[key]['first_invoice'] is None:
            total_data[key]['first_invoice'] = row
    
    # ساخت سطرهای جدید برای حالت total به تفکیک تاریخ
    total_rows = []
    
    # مرتب‌سازی بر اساس تاریخ و شماره بانک
    sorted_keys = sorted(total_data.keys(), key=lambda x: (x[0], x[1]))
    
    base_row_number = create_base_row_number()
    base_row_number = 10000
    iter__ = 0 
    for payment_date, bank_code in sorted_keys:
        values = total_data[(payment_date, bank_code)]
        iter__+=1
        # محاسبه مجموع کل مبلغ دريافت (نقدی + کارت به کارت)
        total_receipt = values['kart_be_kart_total']
        number = f'{base_row_number}00{iter__}'
        # ایجاد سطر تجمعی برای هر تاریخ و بانک
        type= 'ReceiptDraft'
        sandogh = ''


        total_row = {
            'نوع قلم': type,
            'رسيد دريافت نوع دريافت': 1,
            'رسيد دريافت کد': MOSHTARAK_DEFAULT_CODE,
            'رسيد دريافت شماره': number,
            'رسيد دريافت تاريخ': payment_date,  # تاریخ بدون ساعت
            'رسيد دريافت كد معين': 121201,
            'رسيد دريافت صندوق': sandogh,
            'رسيد دريافت مبلغ نقد': 0,
            'رسيد دريافت شرح': f'بابت فروش {selected_date}',
            'رسيد دريافت مبلغ دريافت': total_receipt,
            'حواله شماره': number,
            'حواله تاريخ': payment_date,  # تاریخ بدون ساعت
            'حواله مبلغ': values['kart_be_kart_total'],
            'حواله شرح': f'بابت فروش {selected_date}',
            'حواله تفصيل حساب بانكي': bank_code,
            'حواله حساب بانکی': bankd_detail[bank_code],
            'رسید دریافت تخفیف': 0,
            'رسید دریافت استقراری': 'False'
        }
        total_rows.append(total_row)



        if values['naghdi_total']>0:
            type = ''
            sandogh = 'صندوق مركزي'

            iter__+=1
            # محاسبه مجموع کل مبلغ دريافت (نقدی + کارت به کارت)
            total_receipt = values['naghdi_total'] 

            number = f'{base_row_number}00{iter__}'



            total_row = {
                'نوع قلم': type,
                'رسيد دريافت نوع دريافت': 1,
                'رسيد دريافت کد': MOSHTARAK_DEFAULT_CODE,
                'رسيد دريافت شماره': number,
                'رسيد دريافت تاريخ': payment_date,  # تاریخ بدون ساعت
                'رسيد دريافت كد معين': 121201,
                'رسيد دريافت صندوق': sandogh,
                'رسيد دريافت مبلغ نقد': values['naghdi_total'],
                'رسيد دريافت شرح': f'مجموع مبالغ نقد بانک {bank_code} - تاریخ {payment_date}',
                'رسيد دريافت مبلغ دريافت': total_receipt,
                # 'حواله شماره': number,
                # 'حواله تاريخ': payment_date,  # تاریخ بدون ساعت
                # 'حواله مبلغ': values['kart_be_kart_total'],
                # 'حواله شرح': f'مجموع کارت به کارت بانک {bank_code} - تاریخ {payment_date}',
                # 'حواله تفصيل حساب بانكي': bank_code,
                # 'حواله حساب بانکی': bankd_detail[bank_code],
                # 'رسید دریافت تخفیف': 0,
                'رسید دریافت استقراری': 'False'
            }
            total_rows.append(total_row)



    for row in other_rows:
        total_rows.append(row)





    return total_rows


def create_base_row_number():
    # Get current time in Jalali (Persian) calendar
    jalali_time = jdatetime.datetime.now()
    
    year = jalali_time.year
    month = f"{jalali_time.month:02d}"  # Format month as 2 digits (01, 02, ..., 12)
    day = f"{jalali_time.day:02d}"      # Format day as 2 digits (01, 02, ..., 31)
    
    return f'{year}{month}{day}'



# views.py
from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from .models import Invoice, InvoiceItem, Payment
import json


from django.contrib.auth.decorators import login_required

@login_required
def factor_list(request):
    """View to display factors for the logged-in user"""
    # Get the user's code_vaset from their profile
    pryk_id = request.user.profile.code_vaset
    
    # Filter invoices where peyk field matches user's code_vaset
    invoices = Invoice.objects.filter(peyk=pryk_id).order_by('-created_at')
    
    # Calculate payment summaries for each invoice
    for invoice in invoices:
        payments = invoice.payments.all()
        invoice.paid_amount = sum(p.amount for p in payments)
        invoice.remaining = invoice.total_price - invoice.paid_amount
        
        # Group payments by method
        payment_methods = {}
        for payment in payments:
            if payment.method not in payment_methods:
                payment_methods[payment.method] = 0
            payment_methods[payment.method] += payment.amount
        
        invoice.payment_methods = payment_methods
    
    # Add count for empty state message
    context = {
        'invoices': invoices,
        'invoices_count': invoices.count(),
        'user_code': pryk_id
    }
    
    return render(request, 'factor_list.html', context)



from django.core.exceptions import PermissionDenied

@login_required
def factor_detail(request, invoice_id):
    """View to show factor details and payment management"""
    pryk_id = request.user.profile.code_vaset
    
    # Get invoice and verify it belongs to this user
    invoice = get_object_or_404(
        Invoice.objects.prefetch_related('items', 'payments'), 
        id=invoice_id,
        peyk=pryk_id  # This ensures the invoice belongs to the logged-in user
    )
    
    # Get existing payments
    payments = invoice.payments.all()
    total_paid = sum(p.amount for p in payments)
    remaining = invoice.total_price - total_paid
    
    # Group payments by method
    payment_by_method = {}
    for payment in payments:
        if payment.method not in payment_by_method:
            payment_by_method[payment.method] = 0
        payment_by_method[payment.method] += payment.amount
    
    context = {
        'invoice': invoice,
        'items': invoice.items.all(),
        'payments': payments,
        'total_paid': total_paid,
        'remaining': remaining,
        'payment_by_method': payment_by_method,
        'payment_methods': Payment.PaymentMethod.choices,
    }
    
    return render(request, 'factor_detail.html', context)






@csrf_exempt
def update_payments(request, invoice_id):
    """API endpoint to update payments"""
    if request.method == 'POST':
        try:
            invoice = get_object_or_404(Invoice, id=invoice_id)
            data = json.loads(request.body)
            
            # Delete existing payments
            invoice.payments.all().delete()
            
            # Create new payments
            for method_data in data['payments']:
                if int(method_data['amount']) > 0:
                    Payment.objects.create(
                        invoice=invoice,
                        method=method_data['method'],
                        amount=int(method_data['amount']),
                        created_at=timezone.now()
                    )
            
            return JsonResponse({
                'success': True,
                'message': 'Payments updated successfully'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=400)
    
    return JsonResponse({'error': 'Invalid method'}, status=405)



from django.http import JsonResponse
from django.core.paginator import Paginator
from django.utils import timezone
from datetime import datetime

@login_required
def api_factors(request):
    """API endpoint for infinite scroll loading"""
    pryk_id = request.user.profile.code_vaset
    page = int(request.GET.get('page', 1))
    filter_type = request.GET.get('filter', 'today')  # 'today' or 'all'
    pryk_id = int(pryk_id)
    # Base queryset
    queryset = Invoice.objects.filter(peyk=pryk_id)
    
    # Apply filter
    if filter_type == 'today':
        today = timezone.now().date()
        queryset = queryset.filter(created_at__date=today)
    
    # Order by date (newest first)
    queryset = queryset.order_by('-created_at')
    
    # Paginate (20 items per page)
    paginator = Paginator(queryset, 20)
    current_page = paginator.get_page(page)
    
    # Prepare data
    invoices_data = []
    for invoice in current_page.object_list:
        payments = invoice.payments.all()
        paid_amount = sum(p.amount for p in payments)
        
        # Group payments by method
        payment_methods = {}
        for payment in payments:
            if payment.method not in payment_methods:
                payment_methods[payment.method] = 0
            payment_methods[payment.method] += payment.amount
        
        invoices_data.append({
            'id': invoice.id,
            'invoice_number': invoice.invoice_number,
            'name': invoice.name,
            'phone': invoice.phone,
            'total_price': float(invoice.total_price),
            'paid_amount': float(paid_amount),
            'remaining': float(invoice.total_price - paid_amount),
            'created_at': invoice.created_at.isoformat(),
            'payment_methods': payment_methods
        })
    
    return JsonResponse({
        'invoices': invoices_data,
        'has_more': current_page.has_next(),
        'current_page': page,
        'total_pages': paginator.num_pages
    })







def all_contancts_excel(request):



    from persian_gender_detection import get_gender,clean_name

    # -----------------------------
    # 1) Build rows (one row per invoice item)
    # -----------------------------
    rows = []
    phone_name_pairs = Invoice.objects.values_list('phone','name').distinct()

    for inv in phone_name_pairs:
        # try:
            if inv[0]!='':
                phone = inv[0]
                name = clean_name(inv[1])

                if 'اسنپ' in name:
                    name = name.replace('اسنپ','')
                if name == 'مشتری' or name=='':
                    continue
                if len(name)<=1:
                    continue
                w = is_persian_name(name)
                # if w:
                # if name!='':
                #     print(w)
                ret = evaluate_phone_number(phone_number=phone)
                if ret:
                    rows.append({
                        'اسم' : name,
                        'همراه':phone,


                    })
        # except :
        #     pass



    # -----------------------------
    # 2) Template + mapping
    # -----------------------------

    for row in rows:
        if row['اسم']=='':
            print('empty')



    # Load once at import time
    SERVER = check_server()
    if SERVER:
        template_path = r"/home/seketal1/Seketala_Kitchen_Flow/cache/contact.xlsx"  # must be .xlsx
    else:
        template_path = r'cache\contact.xlsx'
    wb = load_workbook(template_path)
    ws = wb[wb.sheetnames[0]]  # or wb["Sheet1"]

    START_ROW =2 # where the first data row begins in your template

    # Only these columns will be filled; everything else stays unchanged
    # مثال: اگر نمیخوای ستون C پر بشه، اصلا اینجا قرارش نده
    COL_MAP = {
            'اسم' : 'A',
            'همراه':'D',


        # "date": "M",   # if you don't want it, comment/remove it
    }

    # Precompute numeric column indexes (faster)
    col_idx_map = {k: column_index_from_string(v) for k, v in COL_MAP.items()}

    # -----------------------------
    # 3) Write only mapped columns
    # -----------------------------
    for i, record in enumerate(rows):
        excel_row = START_ROW + i
        for field, col_idx in col_idx_map.items():
            ws.cell(row=excel_row, column=col_idx, value=record.get(field))

    # -----------------------------
    # 4) Return as download
    # -----------------------------
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    output = convert_xlsx2xls(output=output)

    j_date = jdatetime.date.today()

    filename = f"resid_sepidar_{j_date}.xls"
    resp = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.ms-excel",  # مقدار جدید برای XLS
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp




import re

def evaluate_phone_number(phone_number):
    """
    یک تابع ساده برای ارزیابی شماره تلفن.
    """
    # حذف کاراکترهای غیر عددی
    phone_number = re.sub(r'\D', '', phone_number)

    # بررسی طول شماره
    if len(phone_number) <= 10 or len(phone_number) > 11:
        return False

    # بررسی قالب شماره (مثال: شروع با 0 یا +98)
    if not (phone_number.startswith('0') or phone_number.startswith('+98')):
        return False

    if phone_number == '09130000000':
        return False

    # بررسی بیشتر (می‌توانید قوانین خاص خود را اضافه کنید)
    # ...

    return True




from persian_names import male_names_fa,female_names_fa

# تبدیل به set برای جستجوی سریع
male_set   = set(male_names_fa)
female_set = set(female_names_fa)

def is_persian_name(word: str) -> bool:
    w = word.strip().lower()
    return w in male_set or w in female_set





def generate_unique_username(base_name):
    """ساخت نام کاربری یکتا با اضافه کردن عدد در صورت تکراری بودن"""
    username = base_name
    counter = 1
    
    while User.objects.filter(username=username).exists():
        username = f"{base_name}{counter}"
        counter += 1
    
    return username





from users.models import Profile, User, jobs
from Constatns import Constants



class ReceiveUser(APIView):
    def post(self, request):
        data = request.data

        # ── احراز هویت ──
        if request.headers.get("X-API-KEY") != "SECRET123":
            return Response({"error": "unauthorized"}, status=403)

        # ── دریافت فیلدها ──
        name     = data.get('name')
        eshterak = data.get('eshterak')
        semat    = data.get('semat')
        sematid  = data.get('sematid')
        phone    = data.get('phone')
        
        print('eshterak'*80)

        # ── اعتبارسنجی فیلدهای اجباری ──
        if not all([name, eshterak, semat, sematid]):
            return Response(
                {"error": "Some Data not Found"}, 
                status=status.HTTP_400_BAD_REQUEST  # 403 → 400 برای bad request بهتر است
            )
        
        if 'متفرقه' in name:

            return Response(
                {"error": "Data not Correct"}, 
                status=status.HTTP_400_BAD_REQUEST  # 403 → 400 برای bad request بهتر است
            )
        

        check_eshterak = Profile.objects.filter(code_vaset=eshterak)
        if check_eshterak.exists():

            return Response(
                {"error": "Eshterak Exist"}, 
                status=status.HTTP_400_BAD_REQUEST  # 403 → 400 برای bad request بهتر است
            )

        # ── جداسازی نام و نام خانوادگی ──
        parts = name.split(' ')
        if len(parts) < 2:
            return Response(
                {"error": "Name must contain first and last name"}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        f_name = parts[0]
        l_name = ' '.join(parts[1:])  # برای نام‌های چندکلمه‌ای
        # ── ساخت نام کاربری یکتا ──
        username = generate_unique_username(f_name)
        # ── ساخت کاربر ──
        password = Constants.DEFAULT_PASSWORD

        user,user_created = User.objects.get_or_create(
            username=username, 
            password=password
        )

        # ── ساخت یا دریافت شغل ──
        short_name = semat[:4] if len(semat) >= 4 else semat
        job, _ = jobs.objects.get_or_create(
            name=semat,
            defaults={
                'persian_name': semat,
                'short_name': short_name
            }
        )
        profile, created = Profile.objects.get_or_create(user=user)
        profile.phone = phone
        profile.first_name = f_name
        profile.last_name = l_name
        profile.job_position = job
        profile.code_vaset = eshterak
        profile.save()



        sms_template = SMS_Template.objects.filter(name =SMSServiceTemplate_Enum.AUTO_SIGNUP )
        if sms_template.exists():
            sms_template = sms_template.first()
            sms_recievers = SMS_Recievers.objects.filter(template = sms_template)
            for sms_rec in sms_recievers:
                send_phone = sms_rec.persons.phone
                f_name = sms_rec.persons.f_name


                ret = send_sms(sms_template,phone_number=send_phone,vars={OTPVar_Enum.NAME:username,OTPVar_Enum.NAME:name,OTPVar_Enum.JOBNAME:semat,OTPVar_Enum.PHONE:phone,})









        if created:
            return Response({
                "status": "ok",
                "username": username          # ✅ نام کاربری برگردانده می‌شود
            }, status=status.HTTP_201_CREATED)
        else:
            return Response({
                "status": "Existed",
                "username": username
            }, status=status.HTTP_200_OK)
        


class ReceiveUpdateInvoiceTime(APIView):
    def post(self, request):
        data = request.data

        # ── Authentication ──
        if request.headers.get("X-API-KEY") != "SECRET123":
            return Response({"error": "unauthorized"}, status=403)

        # ── Get fields ──
        invoice_number = data.get('invoice_number')
        date = data.get('date')  # Jalali date: e.g., "1404/07/15"
        time = data.get('time')  # Time: e.g., "14:30:00"

        # ── Validate required fields ──
        if not invoice_number:
            return Response({"error": "invoice_number is required"}, status=status.HTTP_400_BAD_REQUEST)
        
        if not date or not time:
            return Response({"error": "date and time are required"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            # ── Convert Jalali date/time to Gregorian ──
            date_time = jalali_date_time_to_gregorian(date,time)
        except Exception as e:
            return Response({"error": f"Invalid date/time format: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

        # ── Create unique invoice number ──
        year_shamsi = jdatetime.datetime.now().year
        # unique_invoice_number = f'{year_shamsi}_{invoice_number}'

        # ── Check if invoice exists ──
        try:
            invoice = Invoice.objects.filter(invoice_number=invoice_number).last()
        except Invoice.DoesNotExist:
            return Response({
                "error": f"Invoice with number {invoice_number} does not exist",
                "message": "Invoice not found"
            }, status=status.HTTP_404_NOT_FOUND)

        # ── Update the time ──
        invoice.created_at = date_time
        invoice.save()

        # ── Return success response ──
        return Response({
            "success": True,
            "message": f"Invoice {invoice_number} updated successfully",
            "data": {
                "invoice_number": invoice_number,
                "unique_invoice_number": invoice_number,
                "new_time": date_time.strftime("%Y-%m-%d %H:%M:%S")
            }
        }, status=status.HTTP_200_OK)
    



        

class ReceiveRemoveInvoice(APIView):
    def post(self, request):
        data = request.data
        print('start remove')
        # ── Authentication ──
        if request.headers.get("X-API-KEY") != "SECRET123":
            return Response({"error": "unauthorized"}, status=status.HTTP_403_FORBIDDEN)

        # ── Get fields ──
        invoice_number = data.get('invoice_number')
        print('Invoice for delete recieved : ',invoice_number)
        # ── Validate required fields ──
        if not invoice_number:
            return Response(
                {"error": "invoice_number is required"}, 
                status=status.HTTP_400_BAD_REQUEST
            )


        # ── Check if invoice exists ──
        try:
            invoice = Invoice.objects.get(invoice_number=invoice_number)
        except Invoice.DoesNotExist:
            return Response({
                "error": f"Invoice with number {invoice_number} does not exist",
                "message": "Invoice not found"
            }, status=status.HTTP_404_NOT_FOUND)

        # ── Delete the invoice ──
        invoice.delete()
        print('removed success')
        # ── Return success response ──
        return Response({
            "success": True,
            "message": f"Invoice {invoice_number} removed successfully",
            "data": {
                "invoice_number": invoice_number,
                "unique_invoice_number": invoice_number,
                "deleted_at": jdatetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }
        }, status=status.HTTP_200_OK)

