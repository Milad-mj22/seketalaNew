import os

import jdatetime
from datetime import datetime
import re
def jalali_date_time_to_gregorian(jalali_date, time_str):
    """
    jalali_date: '1404-10-12'
    time_str: '14:32:10'
    """
    jy, jm, jd = map(int, jalali_date.split("-"))
    hour, minute = map(int, time_str.split(":"))

    j_date = jdatetime.date(jy, jm, jd)
    g_date = j_date.togregorian()

    return datetime(
        g_date.year,
        g_date.month,
        g_date.day,
        hour,
        minute,
       
    )



import re

def extract_payment_methods(nahveh_text):
    """
    استخراج روش پرداخت از رشته nahveh
    - اگر [] وجود داشته باشد، محتوا داخل براکت برگردد
    - اگر [] نبود، اگر شامل 'نقدي' بود فقط 'نقدي' برگردد
    - در غیر اینصورت کل متن را برگرداند
    """
    if not nahveh_text:
        return set()

    # 1️⃣ اگر براکت وجود داشته باشد
    if '[' in nahveh_text and ']' in nahveh_text:
        matches = re.findall(r'\[\s*(.*?)\s*\]', nahveh_text)
        return set(matches) if matches else set()
    
    # 2️⃣ اگر شامل "نقدي" بود
    elif 'نقدي' in nahveh_text:
        return {'نقدي'}
    
    # 3️⃣ در غیر اینصورت کل متن
    else:
        return {nahveh_text.strip()}



from datetime import datetime, time, timedelta
from django.utils import timezone

from datetime import datetime, time, timedelta
from django.utils import timezone


def get_date_range(selected_date):
    # Start at 03:00 of selected date
    start = datetime.combine(selected_date, time(3, 0))

    # End at 03:00 of next day
    end = start + timedelta(days=1)

    return start, end



def get_date_range_night_form(selected_date):
    # Start at 03:00 of selected date
    start = datetime.combine(selected_date, time(3, 0))

    # End at 03:00 of next day
    end = start + timedelta(days=+1)

    return start, end


from user_management.utils import check_server
SERVER = check_server()
from openpyxl import load_workbook


def read_excel(excel_name:str):

    if SERVER:
        template_path = f"/home/seketal1/Seketala_Kitchen_Flow/cache/{excel_name}"  # must be .xlsx
    else:
        template_path = f'cache\{excel_name}'


        # بارگذاری فایل و خواندن داده‌ها
    wb = load_workbook(template_path)
    ws =  wb[wb.sheetnames[0]] # یا wb["نام_سیت"] برای انتخاب سیت خاص

    # خواندن تمام داده‌های سیت
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(list(row))

    return data



def get_persian_date_string(date_obj):
    """تبدیل تاریخ میلادی به رشته شمسی فارسی"""
    import jdatetime

    current_time = timezone.now()

    if current_time.hour <= 3:
        date_obj-= timedelta(days=1)
    
    persian_date = jdatetime.date.fromgregorian(date=date_obj)
    


    days_name = [ 'شنبه', 'یکشنبه','دوشنبه', 'سه‌شنبه', 'چهارشنبه', 'پنجشنبه', 'جمعه']
    months_name = ['', 'فروردین', 'اردیبهشت', 'خرداد', 'تیر', 'مرداد', 'شهریور', 
                   'مهر', 'آبان', 'آذر', 'دی', 'بهمن', 'اسفند']
    
    persian_digits = '۰۱۲۳۴۵۶۷۸۹'
    
    def to_persian(num):
        return ''.join(persian_digits[int(d)] for d in str(num))
    
    return f"{days_name[persian_date.weekday()]} {to_persian(persian_date.day)} {months_name[persian_date.month]} {to_persian(persian_date.year)}"




import re

def normalize_iranian_mobile(phone):
   
    clear_phone_number = []
   
    try:
        phone = str(phone)
    except Exception as e:
        return False,clear_phone_number
    phone = re.sub(r'\D', '', phone)


    if phone.startswith('0098'):
        phone = phone[4:]
    elif phone.startswith('98'):
        phone = phone[2:]
    elif phone.startswith('0'):
        phone = phone[1:]

    if re.fullmatch(r'9\d{9}', phone):
        clear_phone_number =  [f'+98{phone}']


    if clear_phone_number:
        if '1111111' not in phone:
            return True , clear_phone_number

    return False , clear_phone_number



def check_personel_noght_order(ret, phones: list, data: dict):
    new_ret = False
    phones = [] if phones is None else phones.copy()  # Preserve existing phones if needed
    
    # Check if NIGHT_SMS_DEVELOPE is set to 'true' (case-insensitive)
    night_sms_develope = os.getenv('NIGHT_SMS_DEVELOPE', '').lower() == 'true'
    
    # Check for Alireza
    if 'عليرضا حلبيان فر ' in data.get('name', ''):
        phone_env = os.getenv('ALIREZA_PHONE')
        if phone_env:
            success, new_phones = normalize_iranian_mobile(phone_env)
            if success and new_phones:
                phones.append(new_phones[0])
                new_ret = True
    
    # Check for Majid
    if 'مجيد حلبيان فر ' in data.get('name', ''):
        phone_env = os.getenv('MAJID_PHONE')
        if phone_env:
            success, new_phones = normalize_iranian_mobile(phone_env)
            if success and new_phones:
                phones.append(new_phones[0])
                new_ret = True
    
    # Check for development mode with Milad's phone
    if night_sms_develope:
        phone_env = os.getenv('MILAD_PHONE')
        if phone_env:
            success, new_phones = normalize_iranian_mobile(phone_env)
            if success and new_phones:
                phones.append(new_phones[0])
                new_ret = True
    
    return (ret or new_ret), phones
    
