from collections import defaultdict
from datetime import datetime, timedelta
import random

from django.shortcuts import get_object_or_404, render
# Create your views here.
from django.http import HttpResponse
# views.py
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect

from DataAnalysis.models import Invoice
from DataAnalysis.utils import get_date_range, get_date_range_night_form, get_persian_date_string
from otp_manager.models import OTPVar_Enum, SMS_Recievers, SMS_Template, SMSServiceTemplate_Enum
from otp_manager.service import send_sms
from user_management.utils import check_server
from users.models import Profile
from .models import CustomForm, FormField, FormSubmission, FormSubmissionData, NightlyFormHistory
from django.contrib.auth.models import User

@login_required
def create_form(request):
    users = User.objects.all()
    if request.method == "POST":
        title = request.POST['title']
        description = request.POST.get('description', '')
        creator_ids = request.POST.getlist('creators')
        submitter_ids = request.POST.getlist('submitters')
        form = CustomForm.objects.create(title=title, description=description, created_by=request.user)
        form.allowed_creators.set(User.objects.filter(id__in=creator_ids))
        form.allowed_submitters.set(User.objects.filter(id__in=submitter_ids))
        return redirect('add_fields', form_id=form.id)
    return render(request, 'create_form.html', {'users': users})





# @login_required
# def create_form(request):
#     if request.method == 'POST':
#         form = DynamicFormCreateForm(request.POST)
#         if form.is_valid():
#             form_instance = form.save(commit=False)
#             form_instance.created_by = request.user
#             form_instance.save()
#             form.save_m2m()
#             return redirect('add_fields', form_id=form_instance.id)
#     else:
#         form = DynamicFormCreateForm()
#     return render(request, 'forms/create_form.html', {'form': form})

@login_required
def close_form(request, form_id):
    form = get_object_or_404(CustomForm, id=form_id, created_by=request.user)
    form.is_closed = True
    form.save()
    return redirect('add_fields', form_id=form.id)









@login_required
def add_fields(request, form_id):
    form = CustomForm.objects.get(id=form_id)
    if request.method == 'POST':
        label = request.POST['label']
        field_type = request.POST['field_type']
        FormField.objects.create(form=form, label=label, field_type=field_type)
    return render(request, 'add_fields.html', {'form': form})


from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt

@login_required
@csrf_exempt
def submit_form(request, form_id):
    form = CustomForm.objects.get(id=form_id)

    if request.user not in form.allowed_submitters.all():
        return JsonResponse({"status": "error", "message": "شما اجازه ارسال این فرم را ندارید."}, status=403)

    if request.method == "POST":
        submission = FormSubmission.objects.create(form=form, submitted_by=request.user)
        for field in form.fields.all():
            value = request.POST.get(str(field.id), '')
            FormSubmissionData.objects.create(submission=submission, field=field, value=value)
        return JsonResponse({"status": "success", "message": "فرم با موفقیت ارسال شد."})

    return render(request, 'submit_form.html', {'form': form})



# views.py
from django.shortcuts import render, get_object_or_404
from .models import CustomForm, FormSubmission, FormField, FormSubmissionData

def form_results(request):
    forms = CustomForm.objects.all()
    selected_form = None
    submissions = None
    selected_submission = None
    answers = None

    form_id = request.GET.get('form')
    submission_id = request.GET.get('submission')

    if form_id:
        selected_form = get_object_or_404(CustomForm, id=form_id)
        submissions = FormSubmission.objects.filter(form=selected_form)

    if submission_id:
        selected_submission = get_object_or_404(FormSubmission, id=submission_id)
        answers = FormSubmissionData.objects.filter(submission=selected_submission)

    context = {
        'forms': forms,
        'selected_form': selected_form,
        'submissions': submissions,
        'selected_submission': selected_submission,
        'answers': answers,
    }
    return render(request, 'form_results.html', context)





# # views.py
# from django.contrib.auth.decorators import login_required
# from django.shortcuts import render

# @login_required
# def available_forms(request):
#     forms = CustomForm.objects.all()

#     return render(request, 'available_forms.html', {'forms': forms})



# views.py
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from .models import CustomForm

@login_required
def available_forms(request):
    all_forms = CustomForm.objects.all()

    # Add `can_submit` flag to each form for this user
    forms_with_permission = []
    for form in all_forms:
        can_submit = request.user in form.allowed_submitters.all()
        forms_with_permission.append({
            'form': form,
            'can_submit': can_submit,
        })

    return render(request, 'available_forms.html', {'forms': forms_with_permission})


from django.db.models import Max
from datetime import datetime, timedelta, time

def calc_pardakht(date = None):



    date_str = date
    current_time = timezone.now()
    if date_str:
        selected_date = datetime.strptime(date_str, "%Y-%m-%d").date()
    else:
        selected_date = timezone.now().date()
        if current_time.hour <= 3:
            selected_date-= timedelta(days=1)
            

    start, end = get_date_range_night_form(selected_date)

    invoices = Invoice.objects.filter(
        created_at__range=(start, end)
    ).prefetch_related("items")
    
    totals = defaultdict(int)

    total_items = 0
    nes_items = 0
    kiosk1,kiosk2,kiosk3 = 15,16,17
    komision_peyk = 10/100
    nesieh = {}

    all_parsian = {
    'parsian_kiosk_1': 8,
    'parsian_kiosk_2': 9,
    'parsian_kiosk_3':10,
    'parsian_pain':12,
    'parsian_shomare_1':15,
    'parsian_shomare_2':16,
    'parsian_shomare_3':17,
    'parsian_shomare_4':18,
    'parsian_shomare_5':19,
    }

    print('len(invoices) : ' ,len(invoices))

    for invoice in invoices:



        invoice.pnum = int(invoice.pnum)
        invoice.mablagh_pos = float(invoice.mablagh_pos)
        invoice.hazine_peyk = float(invoice.hazine_peyk)
        invoice.discount = float(invoice.discount)

        if invoice.name == 'کنسل' or invoice.name =='کنسلی':
            continue
        if invoice.items.count() ==1:
            item = invoice.items.all()[0]
            if item.food_name  == '65' and item.quantity ==1:
                continue

        if 'کیوسک۱' in invoice.name or invoice.pnum==kiosk1:
            totals['کیوسک۱'] += invoice.mablagh_pos
        elif 'کیوسک۲' in invoice.name or invoice.pnum==kiosk2:
            totals['کیوسک۲'] +=  invoice.mablagh_pos
        elif 'کیوسک۳' in invoice.name or invoice.pnum==kiosk3:
            totals['کیوسک۳'] +=  invoice.mablagh_pos




        elif 'اسنپ' in invoice.name or invoice.pnum==1:
            totals['اسنپ'] += invoice.total_price - float(invoice.discount) 
            totals['اسنپ پیک'] += float(invoice.hazine_peyk)

        elif float(invoice.mandeh) > 0:
            if float(invoice.hazine_peyk)==0:
                if 20000>float(invoice.moshtarak)>10000 :  ### check is personel
                    nes_items +=1
                    totals['نسیه پرسنل'] += float(invoice.mandeh)

                    nesieh.update({invoice.name:(invoice.total_price +  invoice.hazine_peyk - invoice.discount)})



        if int(invoice.shomare_pos) in all_parsian.values():
            for pos in all_parsian:
                if all_parsian[pos] == int(invoice.shomare_pos):
                    totals[pos] += invoice.mablagh_pos
                    pass





        if  int(invoice.shomare_pos) in all_parsian :
            totals['پارسیان'] += invoice.mablagh_pos 

        if 'واريز به کارت ملي 1]:' in invoice.nahveh or 'واريز به حساب ملي 1]:' in invoice.nahveh:
            totals['واریز1'] +=  float(invoice.nonaghdi)
        elif 'کارت ملي ' in invoice.nahveh or 'حساب ملي ' in invoice.nahveh:
            totals['واریز'] += float(invoice.nonaghdi)

        if float(invoice.discount) > 0:
            totals['تخفیفات'] += float(invoice.discount)


        
        if 'متفرقه' in invoice.nahveh:
            totals['ملی'] += float(invoice.mablagh_pos)


        totals['جمع خالص'] += invoice.total_price + float(invoice.hazine_peyk) 
        totals['نقدی'] += float(invoice.naghdi)
        


        totals['پیک'] += float(invoice.hazine_peyk)


    # print('peyk : ',totals['پیک'])

    totals['مهر'] = 0
    totals['کمیسیون پیک'] = round(totals['پیک']*komision_peyk ,1)



    return totals , nesieh


from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from .forms import NightlySalesForm, save_with_persian_labels
from .models import NightlyFormModel
from decimal import Decimal  # اضافه کردن این خط



from django.utils import timezone



# تابع کمکی برای تبدیل Decimal به float
def convert_decimals_to_floats(data):
    """تبدیل تمام مقادیر Decimal در دیکشنری به float"""
    if isinstance(data, dict):
        for key, value in data.items():
            if isinstance(value, Decimal):
                data[key] = float(value)
            elif isinstance(value, dict):
                convert_decimals_to_floats(value)
    return data

@login_required
def nightly_sales_view(request):
    if request.method == 'POST':
        form = NightlySalesForm(request.POST)
        current_time = timezone.now()
        selected_date = current_time.date()
        persian_date_str = get_persian_date_string(selected_date)
        
        if form.is_valid():
            # تبدیل Decimal به float قبل از ذخیره
            date = request.POST.get('date',False)
            pardakht , nesieh = calc_pardakht(date = None)

            additional_form_dict = get_data_from_form(request=request)
            merged_dict = {**form.cleaned_data,**additional_form_dict}
            cleaned_data = convert_decimals_to_floats(merged_dict)
            cleaned_data = save_with_persian_labels(cleaned_data)
            voice = request.FILES.get('voice_note',None)
            NightlyFormModel.objects.create(
                user=request.user,
                data=cleaned_data,
                voice_note = voice
            )
            
            sms_template = SMS_Template.objects.filter(name =SMSServiceTemplate_Enum.CLOSESANDOGH )
            if sms_template.exists():
                similarity = f"{random.randint(70, 85)}%"
                sms_template = sms_template.first()
                sms_recievers = SMS_Recievers.objects.filter(template = sms_template)
                for sms_rec in sms_recievers:
                    phone = sms_rec.persons.phone
                    f_name = sms_rec.persons.f_name

                    total = 0
                    try:
                        total = cleaned_data['جمع خالص']
                        total = format_number(number=total)
                    except:
                        pass

                    name = 'نامشخص'
                    try:
                        name = request.user.profile.first_name
                    except:
                        pass

                    ret = send_sms(sms_template,phone_number=phone,vars={OTPVar_Enum.NAME:f_name,OTPVar_Enum.CLOSE:name,OTPVar_Enum.DATE:persian_date_str,OTPVar_Enum.VALUE:total,OTPVar_Enum.AMOUNT:similarity,})


            return redirect('success_page')

        return redirect('error_page')
    else:
        pardakht , nesieh = calc_pardakht()
        form = NightlySalesForm(pardakht_data=pardakht)


    context = {
        'form': form,
        'pardakht': pardakht,
        'nesieh': json.dumps(nesieh, ensure_ascii=False),  # ✅ تبدیل به JSON
        'names_list': json.dumps(['milad','ali']),
        # ... سایر داده‌ها
    }

        
    return render(request, 'nightly_form.html',context)



def get_data_from_form(request):
    form_data = request.POST
        
    # داده‌های فرم‌های پویا (فرم 1)
    additional_form_dict = {}
    additional_form_dict.update({'date':request.POST.get('date','')})

    for i in range(1, 30):  # حداکثر 10 فرم
        name = request.POST.get(f'name_{i}')
        value1 = request.POST.get(f'value1_{i}')
        value2 = request.POST.get(f'value2_{i}')
        value3 = request.POST.get(f'value3_{i}')
        value4 = request.POST.get(f'value4_{i}')
        value5 = request.POST.get(f'value5_{i}')
        value6 = request.POST.get(f'value6_{i}')
        value7 = request.POST.get(f'value7_{i}')
        
        if name or value1 or value2 or value3 or value4 or value5 or value6 or value7:
            additional_form_dict.update({f'نام پیک_{i}': name})
            additional_form_dict.update({f'اسنپ_{i}': value1})
            additional_form_dict.update({f'تلفنی_{i}': value2})
            additional_form_dict.update({f'جمع کارکرد_{i}': value3})
            additional_form_dict.update({f'کمیسیون_{i}': value4})
            additional_form_dict.update({f'غذا_{i}': value5})
            additional_form_dict.update({f'انعام_{i}': value6})
            additional_form_dict.update({f'خالص پرداخت_{i}': value7})
    

    for i in range(1, 30):
        f2_value1 = request.POST.get(f'f2_value1_{i}')
        f2_value2 = request.POST.get(f'f2_value2_{i}')

        
        if f2_value1 or f2_value2 :
            additional_form_dict.update({f'شرح_{i}': f2_value1})
            additional_form_dict.update({f'مبلغ - ریال_{i}': f2_value2})

    return additional_form_dict


@login_required
def form_detail(request, form_id):
    form_instance = get_object_or_404(NightlyFormModel, id=form_id, user=request.user)
    now = timezone.now()
    time_diff = now - form_instance.created_at
    is_editable = time_diff.total_seconds() < 7200 or request.user.is_staff

    if request.method == 'POST' and is_editable:
        form = NightlySalesForm(request.POST)
        if form.is_valid():
            # تبدیل Decimal به float قبل از ذخیره
            cleaned_data = convert_decimals_to_floats(form.cleaned_data)
            form_instance.data = cleaned_data
            form_instance.save()
            
            NightlyFormHistory.objects.create(
                form=form_instance,
                user=request.user,
                old_data=form_instance.data,
                new_data=cleaned_data
            )
            return redirect('form_detail', form_id=form_id)
    else:
        # تبدیل Decimal به float برای نمایش در فرم
        initial_data = convert_decimals_to_floats(form_instance.data)
        form = NightlySalesForm(initial=initial_data)

    return render(request, 'form_detail.html', {
        'form': form,
        'form_instance': form_instance,
        'is_editable': is_editable
    })







from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse
from django.views.generic import ListView
from django.urls import reverse_lazy
from .models import NightlyFormModel
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from io import BytesIO
import json
from openpyxl.utils import column_index_from_string

class NightlyFormListView(ListView):
    model = NightlyFormModel
    template_name = 'nightly_forms_list.html'
    context_object_name = 'forms'
    ordering = ['-created_at']
    paginate_by = 30

    def get_queryset(self):
        return NightlyFormModel.objects.order_by('-created_at')[:30]

def download_excel(request, form_id):
    form = get_object_or_404(NightlyFormModel, id=form_id)
    
    # ایجاد کتابکار Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "فرم شبانه"

    new_version = True
    if new_version:
    # اضافه کردن سرستون‌ها
        headers = list(form.data.keys())
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=col_num, column=1, value=header)
            ws.column_dimensions[get_column_letter(col_num)].width = 20
        
        # اضافه کردن داده‌ها
        for col_num, value in enumerate(form.data.values(), 1):
            ws.cell(row=col_num, column=2, value=value)


    else:

        # Load once at import time
        SERVER = check_server()
        if SERVER:
            template_path = r"/home/seketal1/Seketala_Kitchen_Flow/cache/sandogh.xlsx"  # must be .xlsx
        else:
            template_path = r'cache\sandogh.xlsx'
        wb = load_workbook(template_path)
        ws = wb["Sheet3"]  # or wb["Sheet1"]

        # تعریف مپینگ
        cell_mapping = {
            'کارتخوان بانک مهر': ('C', 4),      # C4
            'کارتخوان بانک پارسیان': ('C', 5),      # C4
            'کارتخوان بانک ملی': ('C', 6),      # C4
            'واریزی های بانک.....' :  ('C', 7),      # C4
            'وجه نقد صندوق' :  ('C', 8),      # 
            'نسیه پرسنل' : ('C',9),
            'اسنپ فود' : ('C',10),
            'پیک اسنپ فود' : ('C',11),
            'تخفیفات' : ('C',12),
            'جمع خالص' : ('C',13),
            'فروش ناخالص' : ('F',4),
            'کمیسیون پیک ها' : ('F',5),
            'پرداختی به آقای شمس' : ('F',6),
            'استرداد به مشتری' : ('F',7),
            'جمع ناخالص' : ('F',8),
            'کسر/اضافه صندوق' : ('F',9),
            'سایر هزینه ها' : ('F',10),
            'توضیحات' : ('F',11),
            'date': ('E',2)
        }

        for iter in range(1,6):
            
            cell_mapping.update({f'نام پیک_{iter}':('G',iter+2)})
            cell_mapping.update({f'اسنپ_{iter}':('H',iter+2)})
            cell_mapping.update({f'تلفنی_{iter}':('I',iter+2)})
            cell_mapping.update({f'جمع کارکرد_{iter}':('J',iter+2)})
            cell_mapping.update({f'کمیسیون_{iter}':('K',iter+2)})
            cell_mapping.update({f'غذا_{iter}':('L',iter+2)})
            cell_mapping.update({f'انعام_{iter}':('M',iter+2)})
            cell_mapping.update({f'خالص پرداخت_{iter}':('N',iter+2)})


        nesieh_cells = [('G','H'),('i','j'),('K','L'),('M','N')]
        row_nesieh = [11,12,13]
        iter=1
        for cels in nesieh_cells:
            for row in row_nesieh:

                cell_mapping.update({f'شرح_{iter}':(cels[0],row)})
                cell_mapping.update({f'مبلغ - ریال_{iter}':(cels[1],row)})
                iter+=1



        # پر کردن سلول‌ها
        for key, (col_letter, row_num) in cell_mapping.items():
            if key == 'واریزی های بانک.....' :
                key = 'واریزی های بانک مارال'
                v = form.data.get(key, 0)
                key = 'واریزی های بانک مارینا'
                new_v = form.data.get(key, 0)
                new_text = 'واریزی های بانک '
                if v>0:
                    new_text += ' - مارال'
                if new_v>0:
                    new_text += ' - مارینا'
                    v+=new_v

                    

                ws.cell(
                    row=row_num,
                    column=column_index_from_string('B'),
                    value=new_text
                )
             
            

            else:
                v = form.data.get(key, '')



            try: 
                v= float(v)
            except:
                pass
            ws.cell(
                row=row_num,
                column=column_index_from_string(col_letter),
                value=v
            )


    # تنظیمات خروجی
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # ایجاد پاسخ دانلود
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=form_{form.id}_{form.date}.xlsx'
    return response





def peyk_create(request):
    return render(request, 'form_peyk.html')



# views.py
from django.http import JsonResponse

def get_people(request):
    people = Profile.objects.all().values('id', 'name')
    
    return JsonResponse(list(people), safe=False)



def format_number(number):
  """یک عدد را دریافت کرده و سه رقم سه رقم با کاما جدا می‌کند."""
  return "{:,}".format(number)
