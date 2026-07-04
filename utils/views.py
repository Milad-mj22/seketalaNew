import base64
import difflib
import os
from django.shortcuts import render

# Create your views here.
from django.shortcuts import render
from pathlib import Path
# Create your views here.
from openpyxl.utils import column_index_from_string
import xlrd
import xlwt
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, permission_required
from .utils import detect_gender, fix_persian_text
import json
import pandas as pd
from django.shortcuts import render
from .forms import CSVUploadForm
from users.models import Buyer, Inventory, MaterialComposition, Warehouse , mother_material , raw_material , mode_raw_materials
from .models import RawMaterialTransfer
from user_management.utils import check_server
SERVER = check_server()


@login_required
def import_buyers_csv(request):
    created_count = 0
    updated_count = 0
    skipped_count = 0
    updated_names = []
    created_names = []
    skipped_names = []
    male_count = 0
    female_count = 0
    if request.method == 'POST':
        form = CSVUploadForm(request.POST, request.FILES)
        if form.is_valid():
            csv_file = request.FILES['file']

            col_first = request.POST.get('col_first')
            col_last = request.POST.get('col_last')
            col_phone = request.POST.get('col_phone')



            try:
                df = pd.read_csv(csv_file)
            except Exception as e:
                return render(request, 'import_csv.html', {
                    'form': form,
                    'error': 'خطا در خواندن فایل CSV: ' + str(e),
                })

            for _, row in df.iterrows():
                national_code = str(row.get('national_code', '')).strip()
                phone_number = str(row.get(col_phone, '')).strip()
                first_name = fix_persian_text(str(row.get(col_first, '')))
                last_name = fix_persian_text(str(row.get(col_last, '')))

                if  not phone_number :
                    skipped_count += 1
                    skipped_names.append(f'{first_name} {last_name}')
                    continue
                if  phone_number  == 'nan':
                    skipped_count += 1
                    skipped_names.append(f'{first_name} {last_name}')
                    continue
                

                if first_name == 'nan':
                    first_name = ''
                if last_name =='nan':
                    last_name = ''


                if last_name =='':
                    temp = first_name.split(' ')
                    if len(temp) > 1:
                        last_name = ' '.join(temp[1:])


                buyer = Buyer.objects.filter(first_name=first_name,last_name=last_name).first()

                if buyer:

                    buyer.phone_number = phone_number
                    buyer.save()
                    updated_count += 1
                    updated_names.append(f'{first_name} {last_name} {phone_number}')
                else:

                    # try:
                    buyer_created = False
                    if first_name !='':
                        gender = detect_gender(name=first_name)
                        if gender is not None:
                            gender = gender.lower()
                            if gender in ['male', 'female']:
                                Buyer.objects.create(
                                    first_name=first_name,
                                    last_name=last_name,
                                    phone_number=phone_number,
                                    gender = gender
                                )

                                if gender =='male':
                                    male_count+=1
                                else:
                                    female_count+=1

                                buyer_created = True

                    if not buyer_created:
                        Buyer.objects.create(
                            first_name=first_name,
                            last_name=last_name,
                            phone_number=phone_number,
                        )



                    created_count += 1
                    created_names.append(f'{first_name} {last_name} {phone_number}')


            return render(request, 'import_result.html', {
                'created': created_count,
                'updated': updated_count,
                'skipped': skipped_count,
                'created_names' : created_names,
                'update_names' : updated_names,
                'skipped_names' : skipped_names,
                'male_count':male_count,
                'female_count' : female_count,
                'not_detected' : abs(female_count-male_count),
            })
    else:
        form = CSVUploadForm()

    return render(request, 'import_csv.html', {'form': form})






@login_required
def import_raw_materials_csv(request):
    created_count = 0
    updated_count = 0
    skipped_count = 0
    created_names = []
    updated_names = []
    skipped_names = []

    if request.method == 'POST':
        form = CSVUploadForm(request.POST, request.FILES)
        if form.is_valid():
            csv_file = request.FILES['file']

            try:
                df = pd.read_csv(csv_file)
            except Exception as e:
                return render(request, 'import_csv.html', {
                    'form': form,
                    'error': 'خطا در خواندن فایل CSV: ' + str(e),
                })
            
            col_id = request.POST.get('col_id')
            col_name = request.POST.get('col_name')
            col_unit = request.POST.get('col_unit')
            col_pattern = request.POST.get('col_pattern')
            col_pattern_code = request.POST.get('col_pattern_code')

            


            mode_obj = mode_raw_materials.objects.filter(name = 'خوراکی').first()


            for _, row in df.iterrows():
                name = str(row.get(col_name, '')).strip()
                describe = str(row.get(col_id, '')).strip()
                unit = str(row.get(col_unit, '')).strip()
                mother_name = str(row.get(col_pattern, '')).strip()
                mother_name_code = str(row.get(col_pattern_code, '')).strip()
                mode_name = mode_obj

                if not name or not describe:
                    skipped_count += 1
                    skipped_names.append(name or 'نام نامشخص')
                    continue
                
                if name =='' or describe =='' or unit =='' or mother_name =='' or mother_name_code =='':
                    print('Item scaped',e)
                    skipped_count += 1
                    skipped_names.append(name or 'نام نامشخص')

                    continue


                mother_object = mother_material.objects.get_or_create(name=mother_name,describe=mother_name_code,mode=mode_name)
                
                raw_material_object = raw_material.objects.get_or_create(name=name,describe=describe,unit = unit,mother = mother_object[0],mode=mode_name)
                        

                created_count += 1
                created_names.append(name)

            return render(request, 'import_result_material.html', {
                'created': created_count,
                'updated': updated_count,
                'skipped': skipped_count,
                'created_names': created_names,
                'update_names': updated_names,
                'skipped_names': skipped_names,
            })
    else:
        form = CSVUploadForm()

    return render(request, 'import_csv_material.html', {'form': form})


def create_new_composition_materail(name , code, unit):
    mother_code = code[:4]
    mother_code  = int(float(mother_code))
    from users.models import  raw_material
    mother_obj = mother_material.objects.filter(describe=mother_code).first()

    raw_material_obj = raw_material.objects.get_or_create(name = name,describe =code ,unit = unit,mother=mother_obj)

    return raw_material_obj
    # MaterialComposition.objects.get_or_create(nam)


@login_required
def import_composition_materials_csv(request):
    created_count = 0
    updated_count = 0
    skipped_count = 0
    created_names = []
    updated_names = []
    skipped_names = []

    if request.method == 'POST':
        form = CSVUploadForm(request.POST, request.FILES)
        if form.is_valid():
            csv_file = request.FILES['file']

            try:
                df = pd.read_csv(csv_file)
            except Exception as e:
                return render(request, 'import_csv.html', {
                    'form': form,
                    'error': 'خطا در خواندن فایل CSV: ' + str(e),
                })
            
            main_name = request.POST.get('col_name')
            main_code = request.POST.get('col_code')
            main_unit = request.POST.get('col_unit')
            sub_name = request.POST.get('col_zir_name')
            sub_code = request.POST.get('col_zir_code')
            sub_unit = request.POST.get('col_zir_unit')
            ratio = request.POST.get('col_ratio')

            item_type = 'نوع قلم'



            for _, row in df.iterrows():
                
                value_item_type = str(row.get(item_type, '')).strip()

                if value_item_type != 'FormulaBomItem':
                    
                    continue

                value_main_name = str(row.get(main_name, '')).strip()
                value_main_code = str(row.get(main_code, '')).strip()
                value_main_unit = str(row.get(main_unit, '')).strip()

                value_sub_name = str(row.get(sub_name, '')).strip()
                value_sub_code = str(row.get(sub_code, '')).strip()
                value_sub_unit = str(row.get(sub_unit, '')).strip()

                value_ratio = float(row.get(ratio, 0))



                try:
                    main_material_object = raw_material.objects.filter(name=value_main_name).first()
                    sub_material_obj = raw_material.objects.filter(name=value_sub_name).first()

                    
                    composition_object = MaterialComposition.objects.get_or_create(main_material=main_material_object,ingredient =sub_material_obj,ratio=value_ratio )

                    print(composition_object)
                    created_count+=1
                    created_names.append(main_material_object.name)

                except Exception as e:
                        print('Error in formoula , ',e)
                        skipped_count+=1
                        continue

            return render(request, 'import_result_material.html', {
                'created': created_count,
                'updated': updated_count,
                'skipped': skipped_count,
                'created_names': created_names,
                'update_names': updated_names,
                'skipped_names': skipped_names,
            })
    else:
        form = CSVUploadForm()

    return render(request, 'import_csv_material_copmosition.html', {'form': form})






from decimal import Decimal, InvalidOperation
from django.contrib import messages
from django.db import transaction
from django.db.models import Sum, F
from django.shortcuts import get_object_or_404, redirect, render


from users.models import mother_material as Mother_material  # your current import


def _to_decimal(value, default=Decimal("0")):
    if value is None:
        return default
    try:
        return Decimal(str(float(value)))
    except (ValueError, TypeError, InvalidOperation):
        return default

def manage_inventory(request):
    warehouses = Warehouse.objects.all().order_by("name")
    mother_materials = Mother_material.objects.all().order_by("name")

    # selected warehouse
    selected_warehouse_id = request.POST.get("warehouse") or request.GET.get("warehouse")
    if not selected_warehouse_id and warehouses.exists():
        selected_warehouse_id = str(warehouses.first().id)

    warehouse = None
    if selected_warehouse_id:
        warehouse = get_object_or_404(Warehouse, id=selected_warehouse_id)

    # current stock per mother material (for the selected warehouse)
    mother_stock = {}
    if warehouse:
        agg = (
            Inventory.objects.filter(warehouse=warehouse)
            .values(mother_id=F("inventory_raw_material__mother_id"))
            .annotate(total=Sum("quantity"))
        )
        mother_stock = {row["mother_id"]: row["total"] or Decimal("0") for row in agg}

    # children lists & counts (for collapse section)
    children = raw_material.objects.filter(mother__in=mother_materials).only("id", "name", "mother_id").order_by("name")
    # group children by mother id
    children_by_mother = {}
    for ch in children:
        children_by_mother.setdefault(ch.mother_id, []).append(ch)

    # counts
    child_counts = {mid: len(lst) for mid, lst in children_by_mother.items()}

    # current stock per child (to show inside collapse)
    child_stock = {}
    if warehouse:
        child_agg = (
            Inventory.objects.filter(warehouse=warehouse)
            .values(rm_id=F("inventory_raw_material_id"))
            .annotate(total=Sum("quantity"))
        )
        child_stock = {row["rm_id"]: row["total"] or Decimal("0") for row in child_agg}

    if request.method == "POST":
        action = request.POST.get("action")

        if action == "add_items":
            if not warehouse:
                messages.error(request, "ابتدا یک انبار انتخاب کنید.")
                return redirect("manage_inventory")

            try:
                with transaction.atomic():
                    for mm in mother_materials:
                        qty_raw = request.POST.get(f"qty_{mm.id}")
                        qty_each_child = _to_decimal(qty_raw, Decimal("0"))
                        if qty_each_child > 0:
                            # add same qty to EACH child under this mother
                            for material in children_by_mother.get(mm.id, []):
                                inv, _ = Inventory.objects.get_or_create(
                                    inventory_raw_material=material,
                                    warehouse=warehouse
                                )
                                inv.add_stock(
                                    qty_each_child,
                                    request.user.profile,
                                    receipt_number=-10
                                )
                messages.success(request, "موجودی‌ها با موفقیت اضافه شدند.")
            except Exception as e:
                messages.error(request, f"خطا: {e}")

        elif action == "reset_zero":
            if not warehouse:
                messages.error(request, "ابتدا یک انبار انتخاب کنید.")
                return redirect("manage_inventory")

            try:
                with transaction.atomic():
                    inventories = Inventory.objects.select_for_update().filter(warehouse=warehouse)
                    for inv in inventories:
                        if inv.quantity and inv.quantity > 0:
                            inv.remove_stock(inv.quantity, request.user.profile)
                messages.success(request, "موجودی‌های انبار انتخاب‌شده با موفقیت صفر شد.")
            except Exception as e:
                messages.error(request, f"خطا در صفر کردن موجودی: {e}")

        return redirect(f"{request.path}?warehouse={selected_warehouse_id}")

    return render(request, "manage_inventory.html", {
        "warehouses": warehouses,
        "mother_materials": mother_materials,
        "mother_stock": mother_stock,
        "selected_warehouse_id": selected_warehouse_id,
        "children_by_mother": children_by_mother,
        "child_counts": child_counts,
        "child_stock": child_stock,
    })










@login_required
def convert_raw_materials2sepidar_format(request):

    error_factors = []

    if request.method == 'POST':
        form = CSVUploadForm(request.POST, request.FILES)
        if form.is_valid():
            csv_file = request.FILES['file']
            filename = csv_file.name.lower()
            ext = os.path.splitext(filename)[1]

            try:
                if ext == ".csv":
                    df = pd.read_csv(csv_file)

                elif ext in [".xlsx", ".xls"]:
                    df = pd.read_excel(csv_file)

                else:
                    return render(request, 'import_csv.html', {
                        'form': form,
                        'error': 'فرمت فایل پشتیبانی نمی‌شود. فقط CSV یا Excel.',
                    })

            except Exception as e:
                return render(request, 'import_csv.html', {
                    'form': form,
                    'error': 'خطا در خواندن فایل: ' + str(e),
                })
                        

            first_row = df.iloc[0].to_dict()  # کل ردیف به صورت dict
            
            date = next((col for col in first_row.keys() if '/' in col), None)
            
            csv_file.seek(0)

            if ext == ".csv":
                df = pd.read_csv(csv_file, skiprows=1)
            else:
                df = pd.read_excel(csv_file, skiprows=1)


            col_id = request.POST.get('col_id')
            col_name = request.POST.get('col_name')
            col_unit = request.POST.get('col_unit')
            col_source = request.POST.get('col_source')
            col_dest = request.POST.get('col_destination')

            col_names = {'id':col_id,'name':col_name,'unit':col_unit,'source':col_source,'dest':col_dest}
            outputs,error_factors = create_excel_for_convert_data(col_names,df,date)
            import zipfile,io
            zip_buffer = io.BytesIO()

            with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
                for filename, output in outputs:
                    # CRITICAL FIX: Reset the pointer to the start of the Excel file
                    output.seek(0) 
                    
                    # Write to zip
                    zf.writestr(filename, output.read())

            # CRITICAL FIX: Reset the ZIP buffer to the start before returning
            zip_buffer.seek(0)

            resp = HttpResponse(
                zip_buffer.getvalue(),
                content_type="application/zip"
            )


            file_base64 = base64.b64encode(zip_buffer.read()).decode()

            return JsonResponse({
                "file": file_base64,
                "filename": "sepidar_output.zip",
                "errors": error_factors
            })



            # ... inside your view, after generating 'outputs' ...

            # # If you want to download only the FIRST file:
            # if outputs:
            #     filename, output = outputs[0]  # Take only the first one
                
            #     resp = HttpResponse(
            #         output.getvalue(),
            #         content_type="application/vnd.ms-excel" # Use excel content type
            #     )
            #     resp["Content-Disposition"] = f'attachment; filename="{filename}"'
            #     return resp
            # else:
            #     return render(request, 'import_csv_transfer.html', {'form': form, 'error': 'فایلی تولید نشد'})

            return resp


    else:
        form = CSVUploadForm()

    return render(request, 'import_csv_transfer.html', {'form': form})



def create_excel_for_convert_data(col_names,df,date):
            total_rows = {}
    
            
            error_factors = [] 
            for index, row in df.iterrows():
                row_number = index + 2
                id_ = str(row.get(col_names['id'], '')).strip()
                name =  str(row.get(col_names['name'], '')).strip()
                source = str(row.get(col_names['source'], '')).strip()
                dest = str(row.get(col_names['dest'], '')).strip()
                count = str(row.get(col_names['unit'], '')).strip()
                

                if not id_ or id_.lower() == 'nan':
                    error_factors.append(f'ردیف {row_number}: کد کالا خالی است.')
                    continue

                try:
                    id_ = int(float(id_))
                except Exception:
                    error_factors.append(f'ردیف {row_number}: کد کالا معتبر نیست: {id_}')
                    continue

                if id_ == 0:
                    error_factors.append(f'ردیف {row_number}: کد کالا صفر است.')
                    continue

                if not name or name.lower() == 'nan':
                    error_factors.append(f'ردیف {row_number}: نام کالا خالی است.')
                    continue


                try:
                    count_value = float(count)
                except Exception:
                    error_factors.append(f'ردیف {row_number}: مقدار کالا معتبر نیست: {count}')
                    continue


                raw_object =  raw_material.objects.filter(name=name).last()
                if raw_object is None:
                    raw_object = raw_material.objects.create(name=name,describe=id_,unit='کیلوگرم')
                
                transfer_obj=None
                if source and dest:
                    transfer_obj = RawMaterialTransfer.objects.get_or_create(material=raw_object,source=int(float(source)),destination=int(float(dest)))
                else:
                    transfer_obj = RawMaterialTransfer.objects.filter(material=raw_object)
                    if transfer_obj.exists():
                        # if transfer_obj
                        transfer_obj = transfer_obj.last()
                    #     transfer_obj = transfer_obj.first()
                    else:
                        error_factors.append(
                            f'ردیف {row_number}: برای کالای "{name}" انتقال انبار تعریف نشده است.'
                        )
                        continue


                if not transfer_obj:
                    print('errror in get object transfer object')
                    pass

                if isinstance(transfer_obj, tuple):
                    transfer_obj = transfer_obj[0]


                try:
                    sourece = int(float(transfer_obj.source))
                    destination = int(float(transfer_obj.destination))
                except Exception:
                    error_factors.append(
                        f'ردیف {row_number}: مبدا یا مقصد انتقال برای کالای "{name}" معتبر نیست.'
                    )
                    continue


                if sourece not in total_rows:
                    total_rows[sourece] = []
                
                if float(count)>0:
                    total_rows[sourece].append({
                        'نوع قلم' : 'InventoryDeliveryItem',
                        "خروج انبار نوع": 4,
                        "خروج انبار شماره": '',
                        'خروج انبار تاريخ': date,
                        'خروج انبار كد انبار': sourece,
                        'قلم خروج انبار كد': transfer_obj.material.describe,
                        "قلم خروج انبار واحد اصلي": count,
                        "قلم خروج انبار كد حساب معين": 121506,
                        "خروج انبار كد انبار مقصد": destination,  
                    })


            total_outputs = []
            for source in total_rows:
                output = export_excel(total_rows[source],'sepidar_inventory_movement_template.xlsx')
                filename = f"inventory_movement_{source}.xls"
                
                total_outputs.append(
                    (filename, output)
                )



            return total_outputs , error_factors






from io import BytesIO
from django.http import HttpResponse, JsonResponse
from openpyxl import load_workbook


def export_excel(data, template_name):
    # Load once at import time
        
    template_path = get_template_path(template_name)
    wb = load_workbook(template_path)
    ws = wb[wb.sheetnames[0]]  # or wb["Sheet1"]

    START_ROW =2 # where the first data row begins in your template

    # Only these columns will be filled; everything else stays unchanged
    # مثال: اگر نمیخوای ستون C پر بشه، اصلا اینجا قرارش نده
    COL_MAP = {
            'نوع قلم' : 'A',
            "خروج انبار نوع": 'B',
            "خروج انبار شماره": 'C',
            'خروج انبار تاريخ': 'D',
            'خروج انبار كد انبار': 'E',
            'قلم خروج انبار كد': 'G',
            "قلم خروج انبار واحد اصلي": 'I',
            "قلم خروج انبار كد حساب معين": 'K',
            "خروج انبار كد انبار مقصد": 'L',  
        }

    # Precompute numeric column indexes (faster)
    col_idx_map = {k: column_index_from_string(v) for k, v in COL_MAP.items()}

    # -----------------------------
    # 3) Write only mapped columns
    # -----------------------------
    for i, record in enumerate(data):
        excel_row = START_ROW + i
        for field, col_idx in col_idx_map.items():
            ws.cell(row=excel_row, column=col_idx, value=record.get(field))

    # -----------------------------
    # 4) Return as download
    # -----------------------------
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    output = convert_xlsx2xls(output=output)


    return output


def get_template_path(template_name):
    if SERVER:
        return f"/home/seketal1/Seketala_Kitchen_Flow/cache/{template_name}"

    return Path("cache") / f"{template_name}"


def convert_xlsx2xls(output):

    wb_xlsx = load_workbook(output)
    
    # ایجاد فایل XLS جدید با xlwt
    wb_xls = xlwt.Workbook()
    
    # کپی تمام صفحات به فرمت XLS
    for sheet_name in wb_xlsx.sheetnames:
        ws_xlsx = wb_xlsx[sheet_name]
        ws_xls = wb_xls.add_sheet(sheet_name)
        
        # کپی داده‌ها از XLSX به XLS
        for row_idx, row in enumerate(ws_xlsx.iter_rows()):
            for col_idx, cell in enumerate(row):
                ws_xls.write(row_idx, col_idx, cell.value)
    
    # ذخیره فایل XLS در BytesIO جدید
    output_xls = BytesIO()
    wb_xls.save(output_xls)
    output_xls.seek(0)
    return output_xls
